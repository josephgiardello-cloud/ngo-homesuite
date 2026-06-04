"""Main web routes, dashboards, and Phase 1 CRUD interfaces."""

import csv
import json
import time
import uuid
from ngo_homesuite.ai.copilot_tools import CopilotToolRegistry
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Optional as TypingOptional

from flask import Blueprint, render_template, redirect, url_for, request, flash, send_file, current_app, Response, session, abort, jsonify
from flask_login import login_required, current_user
from flask_wtf import FlaskForm
from werkzeug.exceptions import NotFound
from werkzeug.utils import secure_filename
from wtforms import BooleanField, DateField, FloatField, HiddenField, SelectField, StringField, SubmitField, TextAreaField
from wtforms.validators import DataRequired, Optional as WTOptional, NumberRange, Email, Length, Regexp
from io import BytesIO
from openpyxl import Workbook, load_workbook

from ngo_homesuite.models.core import (
    Organization, Beneficiary, Project, Donation, Donor, Fund, Expense, DonationReceipt, P2PPage, Volunteer, Campaign, EventDiscountCode, Task, CampaignEmailDelivery, DonorEngagementScore, db
)
from ngo_homesuite.services.beneficiary_service import create_beneficiary, get_beneficiary, list_beneficiaries, update_beneficiary
from ngo_homesuite.services.program_impact_service import list_cases
from ngo_homesuite.services.donation_service import DonationConcurrencyError, DonationNotFound, DonationService, InvalidStatusTransition
from ngo_homesuite.services.donor_service import DonorNotFound, DonorService
from ngo_homesuite.services.expense_service import ExpenseService
from ngo_homesuite.services.payment_service import PaymentService, StripeNotConfigured
from ngo_homesuite.services.fund_service import FundConcurrencyError, FundNotFound, FundService
from ngo_homesuite.services.organization_service import get_first_active_organization
from ngo_homesuite.services.project_service import ProjectNotFound, ProjectService
from ngo_homesuite.services.reporting_service import ReportingService
from ngo_homesuite.services.volunteer_service import create_volunteer, list_recent_volunteers
from ngo_homesuite.domain import (
    BeneficiaryEntity,
    CampaignEntity,
    DomainRegistry,
    DonorEntity,
    GrantEntity,
    LifecycleState,
    OutcomeEntity,
    ProgramEntity,
)
from ngo_homesuite.ai.semantic_memory import SemanticMemoryLayer
from ngo_homesuite.services.opinionated_workflows import (
    run_donation_receipt_followup_workflow,
    run_grant_tracking_reporting_workflow,
    run_program_tracking_impact_workflow,
)
from sqlalchemy import func, select
from ngo_homesuite.web.auth_routes import require_step_up_auth
from ngo_homesuite.web import main_api_docs_handlers
from ngo_homesuite.web import main_workflow_handlers
from ngo_homesuite.web.rbac import roles_required
from ngo_homesuite.utils.receipt_pdf import generate_receipt_pdf_bytes
from ngo_homesuite.compliance.evidence_pack import build_compliance_evidence

main_bp = Blueprint('main', __name__)

_SUPPORTED_LOCALES = {'en', 'es', 'fr'}
_DONOR_IMPORT_ALLOWED_EXTENSIONS = {'.csv', '.xlsx'}
_DONOR_IMPORT_MAX_ROWS = 2500
_DONOR_IMPORT_VALID_TYPES = {'individual', 'corporate', 'foundation', 'anonymous'}
_PHOTO_ALLOWED_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}
_PHOTO_MAX_BYTES = 5 * 1024 * 1024
_PUBLIC_DONATION_MAX_ATTEMPTS_PER_HOUR = 3
_UI_PROFILE_MAX_ITEMS = 10
_UI_PROFILE_DEFAULT = {
    "sidebar_collapsed_groups": {},
    "favorites": [],
    "recent": [],
}

# Track process start time for uptime calculation.
_PROCESS_START = time.monotonic()


def _normalize_ui_nav_item(item: object) -> dict[str, str] | None:
    if not isinstance(item, dict):
        return None
    href = str(item.get("href") or "").strip()
    if not href.startswith("/"):
        return None
    label = str(item.get("label") or "").strip()[:80]
    icon = str(item.get("icon") or "").strip()[:8]
    if not label:
        return None
    return {
        "href": href,
        "label": label,
        "icon": icon,
    }


def _normalize_ui_profile(raw: object) -> dict[str, object]:
    payload = raw if isinstance(raw, dict) else {}

    collapsed_raw = payload.get("sidebar_collapsed_groups")
    collapsed: dict[str, bool] = {}
    if isinstance(collapsed_raw, dict):
        for key, value in collapsed_raw.items():
            group = str(key or "").strip().lower()
            if not group:
                continue
            if len(group) > 40:
                continue
            collapsed[group] = bool(value)

    favorites: list[dict[str, str]] = []
    for item in list(payload.get("favorites") or []):
        normalized = _normalize_ui_nav_item(item)
        if normalized and normalized["href"] not in {entry["href"] for entry in favorites}:
            favorites.append(normalized)
            if len(favorites) >= _UI_PROFILE_MAX_ITEMS:
                break

    recent: list[dict[str, str]] = []
    for item in list(payload.get("recent") or []):
        normalized = _normalize_ui_nav_item(item)
        if normalized and normalized["href"] not in {entry["href"] for entry in recent}:
            recent.append(normalized)
            if len(recent) >= _UI_PROFILE_MAX_ITEMS:
                break

    return {
        "sidebar_collapsed_groups": collapsed,
        "favorites": favorites,
        "recent": recent,
    }


def _build_operational_urgency(org_id: int) -> dict[str, int]:
    from ngo_homesuite.services.task_service import overdue_task_summary
    from ngo_homesuite.compliance.monitoring import ComplianceMonitoringService
    from ngo_homesuite.grants.models import GrantApprovalRequest

    try:
        tasks_summary = overdue_task_summary(org_id)
        task_count = int(tasks_summary.get("total_overdue", 0) or 0)
    except Exception:
        task_count = 0

    approvals_count = 0
    if current_user.has_role("admin", "staff"):
        approvals_count = int(
            db.session.scalar(
                select(func.count(GrantApprovalRequest.id)).where(
                    GrantApprovalRequest.organization_id == org_id,
                    GrantApprovalRequest.status.in_(["pending", "escalated"]),
                )
            )
            or 0
        )

    try:
        alerts_payload = ComplianceMonitoringService.check_grant_deadlines(org_id)
    except Exception:
        alerts_payload = {}
    alerts_count = 0
    if isinstance(alerts_payload, dict):
        alerts_count = (
            len(list(alerts_payload.get("critical") or []))
            + len(list(alerts_payload.get("urgent") or []))
            + len(list(alerts_payload.get("warning") or []))
        )

    return {
        "tasks": max(0, task_count),
        "approvals": max(0, approvals_count),
        "alerts": max(0, alerts_count),
    }


def _resolve_event_discount_code(event_id: int, code: str) -> EventDiscountCode | None:
    normalized = (code or "").strip().upper()
    if not normalized:
        return None
    row = db.session.scalars(
        select(EventDiscountCode).where(
            EventDiscountCode.event_id == int(event_id),
            func.upper(EventDiscountCode.code) == normalized,
            EventDiscountCode.is_active.is_(True),
        ).limit(1)
    ).first()
    if row is None:
        return None

    now = datetime.now(timezone.utc).replace(tzinfo=None)
    if row.expires_at and row.expires_at < now:
        return None
    if row.usage_limit is not None and int(row.usage_count or 0) >= int(row.usage_limit):
        return None
    return row


def _discount_amount(base_amount: float, discount: EventDiscountCode | None) -> float:
    if discount is None:
        return 0.0
    dtype = str(discount.discount_type or "").strip().lower()
    value = float(discount.discount_value or 0.0)
    if value <= 0:
        return 0.0
    if dtype == "percentage":
        return round(max(0.0, min(base_amount, (base_amount * value) / 100.0)), 2)
    return round(max(0.0, min(base_amount, value)), 2)


def _public_donation_attempts_last_hour(*, org_id: int, donor_email: str) -> int:
    cutoff = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(hours=1)
    count = db.session.scalar(
        select(func.count(Donation.id)).where(
            Donation.organization_id == int(org_id),
            func.lower(Donation.donor_email) == str(donor_email or "").strip().lower(),
            Donation.payment_method == 'stripe',
            Donation.purpose == 'Public Stripe donation',
            Donation.created_at >= cutoff,
        )
    )
    return int(count or 0)


@main_bp.route('/health', methods=['GET'])
def health() -> Response:
    """Structured health probe: DB reachability, migration version, uptime."""
    from ngo_homesuite.models.core import db as _db

    # DB reachability
    db_ok = False
    db_error: str | None = None
    try:
        _db.session.scalar(_db.text("SELECT 1"))
        db_ok = True
    except Exception as exc:  # pragma: no cover
        db_error = str(exc)

    # Latest applied migration version (from schema_version table)
    migration_version: int | None = None
    try:
        if db_ok:
            raw = _db.session.scalar(_db.text("SELECT MAX(version) FROM schema_version"))
            migration_version = int(raw) if raw is not None else None
    except Exception:  # pragma: no cover
        pass

    # Expected version = highest numbered SQL migration file on disk
    from pathlib import Path as _Path
    _migrations_dir = _Path(__file__).parent.parent / "migrations"
    _sql_files = sorted(_migrations_dir.glob("[0-9]*.sql"))
    expected_version: int = 0
    if _sql_files:
        try:
            expected_version = int(_sql_files[-1].name.split("_")[0])
        except ValueError:  # pragma: no cover
            expected_version = 0

    payload: dict = {
        "status": "ok" if db_ok else "degraded",
        "db": "ok" if db_ok else f"error: {db_error}",
        "migration_version": migration_version,
        "expected_migration_version": expected_version,
        "migration_current": migration_version == expected_version,
        "uptime_seconds": round(time.monotonic() - _PROCESS_START, 1),
    }
    status_code = 200 if db_ok else 503
    return Response(
        json.dumps(payload),
        status=status_code,
        mimetype="application/json",
    )


@main_bp.route('/health/live', methods=['GET'])
def health_live() -> Response:
    """Kubernetes liveness probe — returns 200 as long as the process is running."""
    return Response(
        json.dumps({"status": "live"}),
        status=200,
        mimetype="application/json",
    )


@main_bp.route('/health/ready', methods=['GET'])
def health_ready() -> Response:
    """Kubernetes readiness probe — 200 only when DB is up and migrations are current."""
    from ngo_homesuite.models.core import db as _db
    from pathlib import Path as _Path

    db_ok = False
    try:
        _db.session.scalar(_db.text("SELECT 1"))
        db_ok = True
    except Exception:  # pragma: no cover
        pass

    migration_version: int | None = None
    try:
        if db_ok:
            raw = _db.session.scalar(_db.text("SELECT MAX(version) FROM schema_version"))
            migration_version = int(raw) if raw is not None else None
    except Exception:  # pragma: no cover
        pass

    _migrations_dir = _Path(__file__).parent.parent / "migrations"
    _sql_files = sorted(_migrations_dir.glob("[0-9]*.sql"))
    expected_version: int = 0
    if _sql_files:
        try:
            expected_version = int(_sql_files[-1].name.split("_")[0])
        except ValueError:  # pragma: no cover
            expected_version = 0

    # Ready only when DB is up and migrations are at the expected version (or no SQL migrations exist)
    migration_current = (migration_version == expected_version) or (expected_version == 0 and migration_version is None)
    ready = db_ok and migration_current
    payload = {
        "status": "ready" if ready else "not_ready",
        "db": "ok" if db_ok else "error",
        "migration_current": migration_current,
    }
    return Response(
        json.dumps(payload),
        status=200 if ready else 503,
        mimetype="application/json",
    )


class DonorForm(FlaskForm):
    name = StringField('Name', validators=[DataRequired()])
    salutation = StringField('Salutation', validators=[WTOptional(), Length(max=50)])
    preferred_name = StringField('Preferred Name', validators=[WTOptional(), Length(max=200)])
    email = StringField('Email', validators=[WTOptional(), Email()])
    phone = StringField('Phone', validators=[WTOptional()])
    donor_type = SelectField(
        'Donor Type',
        choices=[
            ('individual', 'Individual'),
            ('corporate', 'Corporate'),
            ('foundation', 'Foundation'),
            ('anonymous', 'Anonymous'),
        ],
        validators=[DataRequired()],
    )
    status = SelectField(
        'Status',
        choices=[
            ('active', 'Active'),
            ('prospect', 'Prospect'),
            ('lapsed', 'Lapsed'),
            ('archived', 'Archived'),
        ],
        validators=[WTOptional()],
    )
    preferred_contact_method = SelectField(
        'Preferred Contact Method',
        choices=[('email', 'Email'), ('phone', 'Phone'), ('mail', 'Mail'), ('none', 'Do Not Contact')],
        validators=[WTOptional()],
    )
    communication_opt_in = BooleanField('Can Contact This Donor', default=True)
    address = StringField('Address', validators=[WTOptional(), Length(max=300)])
    city = StringField('City', validators=[WTOptional(), Length(max=100)])
    country = StringField('Country', validators=[WTOptional(), Length(max=100)])
    postal_code = StringField('Postal Code', validators=[WTOptional(), Length(max=20)])
    employer = StringField('Employer / Organization', validators=[WTOptional(), Length(max=200)])
    source = StringField('Lead Source', validators=[WTOptional(), Length(max=120)])
    notes = TextAreaField('Notes', validators=[WTOptional()])
    submit = SubmitField('Save Donor')


class DonationForm(FlaskForm):
    donor_id = SelectField('Donor', coerce=int, validators=[DataRequired()])
    campaign_id = SelectField('Campaign', coerce=int, validators=[WTOptional()])
    project_id = SelectField('Project', coerce=int, validators=[WTOptional()])
    fund_id = SelectField('Fund', coerce=int, validators=[WTOptional()])
    amount = FloatField('Amount', validators=[DataRequired(), NumberRange(min=0.01)])
    currency = SelectField('Currency', choices=[('USD', 'USD'), ('EUR', 'EUR'), ('GBP', 'GBP')], validators=[DataRequired()])
    payment_method = SelectField(
        'Payment Method',
        choices=[('cash', 'Cash'), ('bank_transfer', 'Bank Transfer'), ('credit_card', 'Credit Card')],
        validators=[DataRequired()],
    )
    channel = SelectField(
        'Channel/Source',
        choices=[
            ('', 'Unspecified'),
            ('web', 'Web'),
            ('event', 'Event'),
            ('mail', 'Mail'),
            ('phone', 'Phone'),
            ('p2p', 'Peer-to-Peer'),
            ('grant_portal', 'Grant Portal'),
        ],
        validators=[WTOptional()],
    )
    is_anonymous = BooleanField('Anonymous/Publicly Hidden')
    public_display_name = StringField('Public Display Name', validators=[WTOptional(), Length(max=200)])
    tribute_type = SelectField(
        'Tribute Type',
        choices=[('', 'None'), ('in_honor_of', 'In Honor Of'), ('in_memory_of', 'In Memory Of')],
        validators=[WTOptional()],
    )
    tribute_honoree_name = StringField('Honoree Name', validators=[WTOptional(), Length(max=200)])
    tribute_honoree_contact = StringField('Honoree Contact', validators=[WTOptional(), Length(max=255)])
    soft_credit_name = StringField('Soft Credit Name', validators=[WTOptional(), Length(max=200)])
    purpose = StringField('Purpose', validators=[WTOptional()])
    reference_number = StringField('Reference Number', validators=[WTOptional()])
    bank_routing_number = StringField('Bank Routing Number', validators=[WTOptional(), Length(max=20)])
    bank_account_number = StringField('Bank Account Number', validators=[WTOptional(), Length(max=32)])
    card_holder_name = StringField('Cardholder Name', validators=[WTOptional(), Length(max=120)])
    card_brand = SelectField(
        'Card Brand',
        choices=[('', 'Select brand'), ('visa', 'Visa'), ('mastercard', 'Mastercard'), ('amex', 'American Express'), ('discover', 'Discover')],
        validators=[WTOptional()],
    )
    card_last4 = StringField('Card Last 4', validators=[WTOptional(), Length(min=4, max=4), Regexp(r'^\d{4}$', message='Card last 4 must be 4 digits')])
    card_exp_month = StringField('Expiry Month', validators=[WTOptional(), Length(min=2, max=2), Regexp(r'^\d{2}$', message='Use MM format')])
    card_exp_year = StringField('Expiry Year', validators=[WTOptional(), Length(min=2, max=4), Regexp(r'^\d{2,4}$', message='Use YY or YYYY format')])
    cash_award_reference = StringField('Cash Award Reference', validators=[WTOptional(), Length(max=100)])
    notes = TextAreaField('Notes', validators=[WTOptional()])
    submit = SubmitField('Record Donation')


class DonorQuickDonationForm(FlaskForm):
    amount = FloatField('Amount', validators=[DataRequired(), NumberRange(min=0.01)])
    currency = SelectField('Currency', choices=[('USD', 'USD'), ('EUR', 'EUR'), ('GBP', 'GBP')], validators=[DataRequired()])
    payment_method = SelectField(
        'Payment Method',
        choices=[('cash', 'Cash'), ('bank_transfer', 'Bank Transfer'), ('credit_card', 'Credit Card')],
        validators=[DataRequired()],
    )
    project_id = SelectField('Project', coerce=int, validators=[WTOptional()])
    fund_id = SelectField('Fund', coerce=int, validators=[WTOptional()])
    purpose = StringField('Purpose', validators=[WTOptional()])
    reference_number = StringField('Reference / Receipt Number', validators=[WTOptional(), Length(max=100)])
    notes = TextAreaField('Notes', validators=[WTOptional()])
    bank_routing_number = StringField('Bank Routing Number', validators=[WTOptional(), Length(max=20)])
    bank_account_number = StringField('Bank Account Number', validators=[WTOptional(), Length(max=32)])
    card_holder_name = StringField('Cardholder Name', validators=[WTOptional(), Length(max=120)])
    card_brand = SelectField(
        'Card Brand',
        choices=[('', 'Select brand'), ('visa', 'Visa'), ('mastercard', 'Mastercard'), ('amex', 'American Express'), ('discover', 'Discover')],
        validators=[WTOptional()],
    )
    card_last4 = StringField('Card Last 4', validators=[WTOptional(), Length(min=4, max=4), Regexp(r'^\d{4}$', message='Card last 4 must be 4 digits')])
    card_exp_month = StringField('Expiry Month', validators=[WTOptional(), Length(min=2, max=2), Regexp(r'^\d{2}$', message='Use MM format')])
    card_exp_year = StringField('Expiry Year', validators=[WTOptional(), Length(min=2, max=4), Regexp(r'^\d{2,4}$', message='Use YY or YYYY format')])
    cash_award_reference = StringField('Cash Award Reference', validators=[WTOptional(), Length(max=100)])
    submit = SubmitField('Record Donation')


class DonorQuickTaskForm(FlaskForm):
    title = StringField('Task Title', validators=[DataRequired(), Length(max=300)])
    task_type = SelectField(
        'Task Type',
        choices=[
            ('follow_up', 'Follow Up'),
            ('call', 'Call'),
            ('email', 'Email'),
            ('meeting', 'Meeting'),
            ('general', 'General'),
        ],
        validators=[DataRequired()],
    )
    priority = SelectField(
        'Priority',
        choices=[('low', 'Low'), ('medium', 'Medium'), ('high', 'High'), ('urgent', 'Urgent')],
        validators=[DataRequired()],
    )
    due_date = DateField('Due Date', validators=[WTOptional()])
    notes = TextAreaField('Task Notes', validators=[WTOptional(), Length(max=2000)])
    submit = SubmitField('Create Task')


class ExpenseForm(FlaskForm):
    project_id = SelectField('Project', coerce=int, validators=[WTOptional()])
    fund_id = SelectField('Fund', coerce=int, validators=[WTOptional()])
    amount = FloatField('Amount', validators=[DataRequired(), NumberRange(min=0.01)])
    currency = SelectField('Currency', choices=[('USD', 'USD'), ('EUR', 'EUR'), ('GBP', 'GBP')], validators=[DataRequired()])
    payee = StringField('Payee', validators=[WTOptional()])
    description = TextAreaField('Description', validators=[WTOptional()])
    submit = SubmitField('Record Expense')


class ProjectForm(FlaskForm):
    name = StringField('Project Name', validators=[DataRequired()])
    description = TextAreaField('Description', validators=[WTOptional()])
    program = StringField('Program', validators=[WTOptional()])
    budget = FloatField('Budget', validators=[DataRequired(), NumberRange(min=0)])
    spent = FloatField('Spent', validators=[DataRequired(), NumberRange(min=0)])
    currency = SelectField('Currency', choices=[('USD', 'USD'), ('EUR', 'EUR'), ('GBP', 'GBP')], validators=[DataRequired()])
    status = SelectField(
        'Status',
        choices=[('planned', 'Planned'), ('active', 'Active'), ('paused', 'Paused'), ('completed', 'Completed')],
        validators=[DataRequired()],
    )
    submit = SubmitField('Save Project')


class FundForm(FlaskForm):
    name = StringField('Fund Name', validators=[DataRequired()])
    description = TextAreaField('Description', validators=[WTOptional()])
    is_active = SelectField(
        'Status',
        choices=[('true', 'Active'), ('false', 'Inactive')],
        validators=[DataRequired()],
    )
    submit = SubmitField('Save Fund')


class ConfirmDeleteForm(FlaskForm):
    submit = SubmitField('Delete')


class FundStatusActionForm(FlaskForm):
    fund_id = HiddenField(validators=[DataRequired()])
    next_url = HiddenField(validators=[WTOptional()])
    set_status = HiddenField(validators=[DataRequired()])
    submit = SubmitField('Update')


class BulkFundActionForm(FlaskForm):
    next_url = HiddenField(validators=[WTOptional()])
    set_status = HiddenField(validators=[DataRequired()])
    submit = SubmitField('Apply')


class DonationRowActionForm(FlaskForm):
    donation_id = HiddenField(validators=[DataRequired()])
    next_url = HiddenField(validators=[WTOptional()])
    new_status = HiddenField(validators=[WTOptional()])
    submit = SubmitField('Submit')


class BulkDonationActionForm(FlaskForm):
    next_url = HiddenField(validators=[WTOptional()])
    new_status = HiddenField(validators=[WTOptional()])
    submit = SubmitField('Apply')


class PublicDonationForm(FlaskForm):
    donor_name = StringField('Full Name', validators=[DataRequired()])
    donor_email = StringField('Email', validators=[WTOptional(), Email()])
    donor_phone = StringField('Phone', validators=[WTOptional()])
    amount = FloatField('Amount', validators=[DataRequired(), NumberRange(min=0.01)])
    currency = SelectField('Currency', choices=[('USD', 'USD'), ('EUR', 'EUR'), ('GBP', 'GBP')], validators=[DataRequired()])
    payment_method = SelectField(
        'Payment Method',
        choices=[('credit_card', 'Credit Card'), ('bank_transfer', 'Bank Transfer'), ('cash', 'Cash')],
        validators=[DataRequired()],
    )
    purpose = StringField('Purpose', validators=[WTOptional()])
    make_recurring = BooleanField('Make this a recurring donation')
    recurring_frequency = SelectField(
        'Recurring Frequency',
        choices=[('monthly', 'Monthly'), ('quarterly', 'Quarterly'), ('yearly', 'Yearly')],
        validators=[WTOptional()],
    )
    submit = SubmitField('Donate')


class P2PPageForm(FlaskForm):
    donor_id = SelectField('Fundraiser Owner', coerce=int, validators=[DataRequired()])
    title = StringField('Fundraiser Title', validators=[DataRequired()])
    goal_amount = FloatField('Goal Amount', validators=[WTOptional(), NumberRange(min=0)])
    story = TextAreaField('Story', validators=[WTOptional()])
    campaign_slug = StringField('Campaign Slug (optional)', validators=[WTOptional()])
    match_ratio = FloatField('Match Ratio (e.g. 1 for 1:1)', validators=[WTOptional(), NumberRange(min=0, max=10)])
    match_cap_amount = FloatField('Match Cap Amount', validators=[WTOptional(), NumberRange(min=0)])
    challenge_goal_amount = FloatField('Challenge Goal Amount', validators=[WTOptional(), NumberRange(min=0)])
    challenge_end_date = DateField('Challenge End Date', validators=[WTOptional()], format='%Y-%m-%d')
    automation_contact_email = StringField('Automation Contact Email', validators=[WTOptional(), Email()])
    submit = SubmitField('Create Fundraiser')


class RecurringDonationForm(FlaskForm):
    donor_id = SelectField('Donor', coerce=int, validators=[DataRequired()])
    amount = FloatField('Amount', validators=[DataRequired(), NumberRange(min=0.01)])
    currency = SelectField('Currency', choices=[('USD', 'USD'), ('EUR', 'EUR'), ('GBP', 'GBP')], validators=[DataRequired()])
    payment_method = SelectField(
        'Payment Method',
        choices=[('credit_card', 'Credit Card'), ('bank_transfer', 'Bank Transfer'), ('cash', 'Cash')],
        validators=[DataRequired()],
    )
    purpose = StringField('Purpose', validators=[WTOptional()])
    frequency = SelectField(
        'Frequency',
        choices=[('monthly', 'Monthly'), ('quarterly', 'Quarterly'), ('yearly', 'Yearly')],
        validators=[DataRequired()],
    )
    submit = SubmitField('Create Recurring Plan')


def _mask_tail(value: str | None, *, keep: int = 4) -> str | None:
    text = str(value or '').strip()
    if not text:
        return None
    if len(text) <= keep:
        return text
    return f"****{text[-keep:]}"


def _build_quick_donation_notes(form: FlaskForm) -> str | None:
    parts: list[str] = []
    notes = str(getattr(form, 'notes', None).data or '').strip() if hasattr(form, 'notes') else ''
    if notes:
        parts.append(notes)

    method = str(getattr(form, 'payment_method', None).data or '').strip().lower()
    if method == 'bank_transfer':
        routing = _mask_tail(getattr(form, 'bank_routing_number', None).data)
        account = _mask_tail(getattr(form, 'bank_account_number', None).data)
        details = ['Method: Bank Transfer']
        if routing:
            details.append(f'Routing: {routing}')
        if account:
            details.append(f'Account: {account}')
        parts.append(' | '.join(details))
    elif method == 'credit_card':
        holder = str(getattr(form, 'card_holder_name', None).data or '').strip()
        brand = str(getattr(form, 'card_brand', None).data or '').strip()
        last4 = _mask_tail(getattr(form, 'card_last4', None).data)
        exp_month = str(getattr(form, 'card_exp_month', None).data or '').strip()
        exp_year = str(getattr(form, 'card_exp_year', None).data or '').strip()
        details = ['Method: Credit Card']
        if holder:
            details.append(f'Cardholder: {holder}')
        if brand:
            details.append(f'Brand: {brand}')
        if last4:
            details.append(f'Last4: {last4}')
        if exp_month or exp_year:
            details.append(f'Expiry: {exp_month}/{exp_year}'.rstrip('/'))
        parts.append(' | '.join(details))
    elif method == 'cash':
        award_ref = str(getattr(form, 'cash_award_reference', None).data or '').strip()
        details = ['Method: Cash']
        if award_ref:
            details.append(f'Award Ref: {award_ref}')
        parts.append(' | '.join(details))

    if not parts:
        return None
    return '\n'.join(parts)


def _build_quick_donation_reference(form: FlaskForm) -> str | None:
    reference = str(getattr(form, 'reference_number', None).data or '').strip() if hasattr(form, 'reference_number') else ''
    if reference:
        return reference
    method = str(getattr(form, 'payment_method', None).data or '').strip().upper() or 'DONATION'
    return f"{method}-{uuid.uuid4().hex[:10].upper()}"


class BeneficiaryIntakeForm(FlaskForm):
    first_name = StringField('First Name', validators=[DataRequired()])
    last_name = StringField('Last Name', validators=[WTOptional()])
    email = StringField('Email', validators=[WTOptional(), Email()])
    phone = StringField('Phone', validators=[WTOptional()])
    date_of_birth = DateField('Date of Birth', validators=[WTOptional()], format='%Y-%m-%d')
    gender = SelectField(
        'Gender',
        choices=[('', '-- Select --'), ('male', 'Male'), ('female', 'Female'), ('non_binary', 'Non-binary'), ('prefer_not_to_say', 'Prefer not to say')],
        validators=[WTOptional()],
    )
    program = StringField('Program / Service Area', validators=[WTOptional()])
    status = SelectField(
        'Status',
        choices=[('active', 'Active'), ('inactive', 'Inactive'), ('pending', 'Pending')],
        validators=[DataRequired()],
    )
    city = StringField('City', validators=[WTOptional()])
    country = StringField('Country', validators=[WTOptional()])
    notes = TextAreaField('Notes', validators=[WTOptional()])
    submit = SubmitField('Save Beneficiary')


def _current_org() -> TypingOptional[Organization]:
    """Pick assigned org first, then fallback to first active org for seeded demo users."""
    if current_user.organization:
        return current_user.organization
    return get_first_active_organization()


def _build_csv_bytes(headers, rows):
    import io
    text_stream = io.StringIO()
    writer = csv.writer(text_stream)
    writer.writerow(headers)
    writer.writerows(rows)
    return text_stream.getvalue().encode('utf-8-sig')


def _build_xlsx_bytes(sheet_name, headers, rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)
    for row in rows:
        worksheet.append(list(row))
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.read()


def _build_iif_bytes(trns_type: str, rows):
    """Build QuickBooks IIF payload bytes for TRNS/SPL/ENDTRNS batches.

    Row schema: [txn_id, date_iso, name, amount, memo]
    """
    text_stream = []
    text_stream.append("!TRNS\tTRNSID\tTRNSTYPE\tDATE\tACCNT\tNAME\tAMOUNT\tMEMO")
    text_stream.append("!SPL\tSPLID\tTRNSTYPE\tDATE\tACCNT\tNAME\tAMOUNT\tMEMO")
    text_stream.append("!ENDTRNS")

    if trns_type == 'DEPOSIT':
        trns_account = 'Undeposited Funds'
        spl_account = 'Donations Income'
        trns_sign = 1.0
    else:
        trns_account = 'Operating Bank'
        spl_account = 'Program Expense'
        trns_sign = -1.0

    for txn_id, date_iso, name, amount, memo in rows:
        try:
            amount_val = float(amount or 0)
        except (TypeError, ValueError):
            amount_val = 0.0
        trns_amt = trns_sign * amount_val
        spl_amt = -trns_amt
        date_value = (date_iso or '').strip() or datetime.now(timezone.utc).strftime('%Y-%m-%d')
        text_stream.append(
            f"TRNS\t{txn_id}\t{trns_type}\t{date_value}\t{trns_account}\t{name}\t{trns_amt:.2f}\t{memo}"
        )
        text_stream.append(
            f"SPL\t{txn_id}\t{trns_type}\t{date_value}\t{spl_account}\t{name}\t{spl_amt:.2f}\t{memo}"
        )
        text_stream.append("ENDTRNS")

    payload = "\n".join(text_stream) + "\n"
    return payload.encode('utf-8')


def _parse_float(value: str):
    value = (value or '').strip()
    if not value:
        return None
    try:
        return float(value)
    except ValueError:
        return None


def _parse_date(value: str):
    value = (value or '').strip()
    if not value:
        return None
    try:
        return datetime.fromisoformat(value).date()
    except ValueError:
        return None


def _normalize_text(value: str) -> str:
    return (value or '').strip().lower()


def _normalize_phone(value: str) -> str:
    return ''.join(ch for ch in (value or '') if ch.isdigit())


def _normalize_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or '').strip().lower() in {'1', 'true', 'yes', 'on', 'y'}


def _donor_import_cache_dir() -> Path:
    cache_dir = Path(current_app.instance_path) / 'donor_import_cache'
    cache_dir.mkdir(parents=True, exist_ok=True)
    return cache_dir


def _uploads_root_dir() -> Path:
    uploads_dir = Path(current_app.instance_path) / 'uploads'
    uploads_dir.mkdir(parents=True, exist_ok=True)
    return uploads_dir


def _save_photo_upload(uploaded, *, entity: str, org_id: int, record_id: int) -> str:
    if uploaded is None or not getattr(uploaded, 'filename', None):
        raise ValueError('No file uploaded')

    filename = secure_filename(str(uploaded.filename or ''))
    if not filename:
        raise ValueError('Invalid file name')

    ext = Path(filename).suffix.lower()
    if ext not in _PHOTO_ALLOWED_EXTENSIONS:
        raise ValueError('Unsupported image type. Allowed: .jpg, .jpeg, .png, .gif, .webp')

    uploaded.stream.seek(0, 2)
    size_bytes = uploaded.stream.tell()
    uploaded.stream.seek(0)
    if size_bytes > _PHOTO_MAX_BYTES:
        raise ValueError('Image must be 5MB or smaller')

    relative_dir = Path(entity) / f'org_{int(org_id)}'
    target_dir = _uploads_root_dir() / relative_dir
    target_dir.mkdir(parents=True, exist_ok=True)
    target_name = f"{int(record_id)}-{uuid.uuid4().hex}{ext}"
    target_path = target_dir / target_name
    uploaded.save(target_path)
    return str((Path('uploads') / relative_dir / target_name).as_posix())


def _resolve_upload_path(relative_path: str) -> Path:
    uploads_root = _uploads_root_dir().resolve()
    candidate = (Path(current_app.instance_path) / relative_path).resolve()
    if uploads_root not in candidate.parents and candidate != uploads_root:
        raise NotFound()
    return candidate


def _parse_donor_import_file(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    ext = path.suffix.lower()
    if ext not in _DONOR_IMPORT_ALLOWED_EXTENSIONS:
        raise ValueError('Only CSV and Excel (.xlsx) files are supported.')

    headers: list[str] = []
    rows: list[dict[str, str]] = []

    if ext == '.csv':
        try:
            with path.open('r', encoding='utf-8-sig', newline='') as fh:
                reader = csv.DictReader(fh)
                headers = [str(h or '').strip() for h in (reader.fieldnames or []) if str(h or '').strip()]
                for idx, row in enumerate(reader, start=1):
                    if idx > _DONOR_IMPORT_MAX_ROWS:
                        break
                    rows.append({h: str((row or {}).get(h, '') or '').strip() for h in headers})
        except UnicodeDecodeError:
            with path.open('r', encoding='latin-1', newline='') as fh:
                reader = csv.DictReader(fh)
                headers = [str(h or '').strip() for h in (reader.fieldnames or []) if str(h or '').strip()]
                for idx, row in enumerate(reader, start=1):
                    if idx > _DONOR_IMPORT_MAX_ROWS:
                        break
                    rows.append({h: str((row or {}).get(h, '') or '').strip() for h in headers})
    else:
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        ws = workbook.active
        sheet_rows = ws.iter_rows(values_only=True)
        raw_headers = next(sheet_rows, None)
        if raw_headers is None:
            return [], []

        headers = [str(c or '').strip() for c in raw_headers]
        headers = [h if h else f'column_{idx + 1}' for idx, h in enumerate(headers)]
        for idx, values in enumerate(sheet_rows, start=1):
            if idx > _DONOR_IMPORT_MAX_ROWS:
                break
            row_map: dict[str, str] = {}
            for col_idx, header in enumerate(headers):
                value = ''
                if values and col_idx < len(values):
                    value = str(values[col_idx] or '').strip()
                row_map[header] = value
            rows.append(row_map)

    return headers, rows


def _guess_donor_import_mapping(headers: list[str]) -> dict[str, str]:
    aliases = {
        'name': {'name', 'full name', 'donor', 'donor name'},
        'salutation': {'salutation', 'title', 'prefix'},
        'preferred_name': {'preferred name', 'preferred_name', 'nickname', 'preferred'},
        'email': {'email', 'email address', 'e-mail'},
        'phone': {'phone', 'phone number', 'mobile', 'telephone'},
        'donor_type': {'type', 'donor type', 'category'},
        'status': {'status', 'crm status', 'donor status'},
        'preferred_contact_method': {'preferred contact method', 'contact preference', 'contact method'},
        'communication_opt_in': {'can contact this donor', 'opt in', 'communication opt in', 'contact ok'},
        'address': {'address', 'street', 'mailing address'},
        'city': {'city', 'town'},
        'country': {'country', 'nation'},
        'postal_code': {'postal code', 'zip', 'zip code', 'postcode'},
        'employer': {'employer', 'organization', 'company'},
        'source': {'source', 'lead source', 'donor source'},
        'notes': {'notes', 'note', 'comments', 'comment'},
    }

    mapping: dict[str, str] = {
        'name': '',
        'salutation': '',
        'preferred_name': '',
        'email': '',
        'phone': '',
        'donor_type': '',
        'status': '',
        'preferred_contact_method': '',
        'communication_opt_in': '',
        'address': '',
        'city': '',
        'country': '',
        'postal_code': '',
        'employer': '',
        'source': '',
        'notes': '',
    }
    normalized = {h: _normalize_text(h) for h in headers}
    for target, candidates in aliases.items():
        for header, normalized_header in normalized.items():
            if normalized_header in candidates and not mapping[target]:
                mapping[target] = header
    if not mapping['name'] and headers:
        mapping['name'] = headers[0]
    return mapping


def _extract_donor_import_mapping(form_data) -> dict[str, str]:
    mapping = {
        'name': (form_data.get('map_name') or '').strip(),
        'salutation': (form_data.get('map_salutation') or '').strip(),
        'preferred_name': (form_data.get('map_preferred_name') or '').strip(),
        'email': (form_data.get('map_email') or '').strip(),
        'phone': (form_data.get('map_phone') or '').strip(),
        'donor_type': (form_data.get('map_donor_type') or '').strip(),
        'status': (form_data.get('map_status') or '').strip(),
        'preferred_contact_method': (form_data.get('map_preferred_contact_method') or '').strip(),
        'communication_opt_in': (form_data.get('map_communication_opt_in') or '').strip(),
        'address': (form_data.get('map_address') or '').strip(),
        'city': (form_data.get('map_city') or '').strip(),
        'country': (form_data.get('map_country') or '').strip(),
        'postal_code': (form_data.get('map_postal_code') or '').strip(),
        'employer': (form_data.get('map_employer') or '').strip(),
        'source': (form_data.get('map_source') or '').strip(),
        'notes': (form_data.get('map_notes') or '').strip(),
    }
    return mapping


def _build_donor_import_preview(org_id: int, rows: list[dict[str, str]], mapping: dict[str, str]) -> dict[str, object]:
    donors = DonorService().list_all_donors(org_id)
    existing_by_email: dict[str, Donor] = {}
    existing_by_name_phone: dict[str, Donor] = {}
    for donor in donors:
        email_key = _normalize_text(donor.email)
        if email_key:
            existing_by_email[email_key] = donor
        name_phone_key = f"{_normalize_text(donor.name)}::{_normalize_phone(donor.phone)}"
        if _normalize_phone(donor.phone):
            existing_by_name_phone[name_phone_key] = donor

    preview_rows: list[dict[str, object]] = []
    error_count = 0
    duplicate_count = 0
    ready_count = 0

    for idx, row in enumerate(rows, start=1):
        mapped_name = str(row.get(mapping['name'], '') if mapping.get('name') else '').strip()
        mapped_salutation = str(row.get(mapping['salutation'], '') if mapping.get('salutation') else '').strip()
        mapped_preferred_name = str(row.get(mapping['preferred_name'], '') if mapping.get('preferred_name') else '').strip()
        mapped_email = str(row.get(mapping['email'], '') if mapping.get('email') else '').strip().lower()
        mapped_phone = str(row.get(mapping['phone'], '') if mapping.get('phone') else '').strip()
        mapped_type = str(row.get(mapping['donor_type'], '') if mapping.get('donor_type') else '').strip().lower() or 'individual'
        mapped_status = str(row.get(mapping['status'], '') if mapping.get('status') else '').strip().lower() or 'active'
        mapped_contact_method = str(row.get(mapping['preferred_contact_method'], '') if mapping.get('preferred_contact_method') else '').strip().lower() or 'email'
        mapped_opt_in = _normalize_bool(row.get(mapping['communication_opt_in'], '') if mapping.get('communication_opt_in') else True)
        mapped_address = str(row.get(mapping['address'], '') if mapping.get('address') else '').strip()
        mapped_city = str(row.get(mapping['city'], '') if mapping.get('city') else '').strip()
        mapped_country = str(row.get(mapping['country'], '') if mapping.get('country') else '').strip()
        mapped_postal_code = str(row.get(mapping['postal_code'], '') if mapping.get('postal_code') else '').strip()
        mapped_employer = str(row.get(mapping['employer'], '') if mapping.get('employer') else '').strip()
        mapped_source = str(row.get(mapping['source'], '') if mapping.get('source') else '').strip()
        mapped_notes = str(row.get(mapping['notes'], '') if mapping.get('notes') else '').strip()

        errors: list[str] = []
        warnings: list[str] = []

        if not mapped_name:
            errors.append('Missing donor name')
        if mapped_type not in _DONOR_IMPORT_VALID_TYPES:
            warnings.append(f"Unknown donor type '{mapped_type}' -> default to individual")
            mapped_type = 'individual'
        if mapped_status not in {'active', 'prospect', 'lapsed', 'archived'}:
            warnings.append(f"Unknown status '{mapped_status}' -> default to active")
            mapped_status = 'active'
        if mapped_contact_method not in {'email', 'phone', 'mail', 'none'}:
            warnings.append(f"Unknown preferred contact method '{mapped_contact_method}' -> default to email")
            mapped_contact_method = 'email'

        duplicate_match = None
        if mapped_email:
            duplicate_match = existing_by_email.get(mapped_email)
        if duplicate_match is None and _normalize_phone(mapped_phone):
            duplicate_match = existing_by_name_phone.get(f"{_normalize_text(mapped_name)}::{_normalize_phone(mapped_phone)}")

        status = 'ready'
        if errors:
            status = 'error'
            error_count += 1
        elif duplicate_match is not None:
            status = 'duplicate'
            duplicate_count += 1
        else:
            ready_count += 1

        preview_rows.append(
            {
                'row_number': idx,
                'name': mapped_name,
                'salutation': mapped_salutation,
                'preferred_name': mapped_preferred_name,
                'email': mapped_email,
                'phone': mapped_phone,
                'donor_type': mapped_type,
                'crm_status': mapped_status,
                'preferred_contact_method': mapped_contact_method,
                'communication_opt_in': mapped_opt_in,
                'address': mapped_address,
                'city': mapped_city,
                'country': mapped_country,
                'postal_code': mapped_postal_code,
                'employer': mapped_employer,
                'source': mapped_source,
                'notes': mapped_notes,
                'import_status': status,
                'errors': errors,
                'warnings': warnings,
                'duplicate_match': duplicate_match,
            }
        )

    return {
        'rows': preview_rows,
        'summary': {
            'total': len(preview_rows),
            'ready': ready_count,
            'duplicates': duplicate_count,
            'errors': error_count,
        },
    }


def _apply_donor_import(org_id: int, preview_rows: list[dict[str, object]]) -> dict[str, int]:
    created = 0
    skipped_duplicates = 0
    skipped_errors = 0
    for item in preview_rows:
        status = str(item.get('status') or '')
        if status == 'error':
            skipped_errors += 1
            continue
        if status == 'duplicate':
            skipped_duplicates += 1
            continue

        DonorService().create_donor(
            org_id,
            str(item.get('name') or '').strip(),
            salutation=(str(item.get('salutation') or '').strip() or None),
            preferred_name=(str(item.get('preferred_name') or '').strip() or None),
            email=(str(item.get('email') or '').strip() or None),
            phone=(str(item.get('phone') or '').strip() or None),
            donor_type=str(item.get('donor_type') or 'individual').strip() or 'individual',
            status=str(item.get('crm_status') or 'active').strip() or 'active',
            preferred_contact_method=str(item.get('preferred_contact_method') or 'email').strip() or 'email',
            communication_opt_in=_normalize_bool(item.get('communication_opt_in', True)),
            address=(str(item.get('address') or '').strip() or None),
            city=(str(item.get('city') or '').strip() or None),
            country=(str(item.get('country') or '').strip() or None),
            postal_code=(str(item.get('postal_code') or '').strip() or None),
            employer=(str(item.get('employer') or '').strip() or None),
            source=(str(item.get('source') or '').strip() or None),
            notes=(str(item.get('notes') or '').strip() or None),
        )
        created += 1

    return {
        'created': created,
        'skipped_duplicates': skipped_duplicates,
        'skipped_errors': skipped_errors,
    }


def _next_charge_date(current: date, frequency: str) -> date:
    if frequency == 'quarterly':
        return current + timedelta(days=90)
    if frequency == 'yearly':
        return current + timedelta(days=365)
    return current + timedelta(days=30)


def _openapi_spec_path() -> Path:
    return Path(current_app.root_path).parent / 'docs' / 'openapi.yaml'


def _sqlite_db_file_path() -> str:
    uri = str(current_app.config.get('SQLALCHEMY_DATABASE_URI', 'sqlite:///ngo_homesuite.db'))
    if uri.startswith('sqlite:///'):
        return uri.replace('sqlite:///', '', 1)
    return 'ngo_homesuite.db'


def _build_domain_registry_for_org(org: Organization | None) -> DomainRegistry:
    registry = DomainRegistry()
    if org is None:
        return registry

    donor_svc = DonorService()
    project_svc = ProjectService()
    reporting_svc = ReportingService()

    donors = donor_svc.list_all_donors(org.id)
    projects = project_svc.list_all_projects(org.id)
    beneficiaries = list_beneficiaries(org.id)

    for donor in donors:
        donor_entity = DonorEntity(
            entity_id=f"donor:{donor.id}",
            name=donor.name,
            donor_type=donor.donor_type,
            email=donor.email,
            phone=donor.phone,
            lifecycle_state=LifecycleState.active,
        )
        donor_entity.transition(LifecycleState.active, actor='system', reason='ingested_for_domain_registry')
        registry.upsert(donor_entity)

    for project in projects:
        program = ProgramEntity(
            entity_id=f"program:{project.id}",
            name=project.name,
            lifecycle_state=LifecycleState.active if project.status in ('active', 'planned') else LifecycleState.paused,
        )
        program.transition(program.lifecycle_state, actor='system', reason='project_to_program_mapping')
        registry.upsert(program)

    for beneficiary in beneficiaries:
        beneficiary_entity = BeneficiaryEntity(
            entity_id=f"beneficiary:{beneficiary.id}",
            name=f"{beneficiary.first_name} {beneficiary.last_name}".strip(),
            lifecycle_state=LifecycleState.active if beneficiary.status == 'active' else LifecycleState.paused,
        )
        beneficiary_entity.transition(beneficiary_entity.lifecycle_state, actor='system', reason='beneficiary_ingested')
        registry.upsert(beneficiary_entity)

    purpose_sums = reporting_svc.donation_purpose_totals(org.id)
    for idx, (purpose, total) in enumerate(purpose_sums, start=1):
        campaign = CampaignEntity(
            entity_id=f"campaign:{idx}",
            name=str(purpose),
            lifecycle_state=LifecycleState.active,
            fundraising_goal=round(float(total) * 1.25, 2),
            raised_amount=float(total),
        )
        campaign.transition(LifecycleState.active, actor='system', reason='derived_from_donation_purpose')
        registry.upsert(campaign)

    foundation_totals = reporting_svc.foundation_donor_totals(org.id)
    for idx, (foundation_name, approved) in enumerate(foundation_totals, start=1):
        grant = GrantEntity(
            entity_id=f"grant:{idx}",
            name=f"{foundation_name} Grant",
            lifecycle_state=LifecycleState.active,
            requested_amount=round(float(approved) * 1.2, 2),
            approved_amount=float(approved),
            status_note='Derived from foundation donor history',
        )
        grant.transition(LifecycleState.active, actor='system', reason='derived_from_foundation_donations')
        registry.upsert(grant)

    project_donation_counts = reporting_svc.project_donation_counts(org.id)
    for project in projects:
        related_donations = project_donation_counts.get(project.id, 0)
        outcome = OutcomeEntity(
            entity_id=f"outcome:program:{project.id}",
            name=f"{project.name} Donor Reach",
            lifecycle_state=LifecycleState.active,
            metric_name='contributing_donations',
            metric_value=float(related_donations),
            program_id=f"program:{project.id}",
        )
        outcome.transition(LifecycleState.active, actor='system', reason='derived_outcome_metric')
        registry.upsert(outcome)
        registry.link(source_id=f"program:{project.id}", relation='has_outcome', target_id=outcome.entity_id, actor='system')

    # Lightweight relationship stitching for usability in semantic retrieval.
    for beneficiary in beneficiaries:
        beneficiary_id = f"beneficiary:{beneficiary.id}"
        for project in projects:
            if beneficiary.program and project.program and beneficiary.program.strip().lower() == project.program.strip().lower():
                registry.link(
                    source_id=beneficiary_id,
                    relation='enrolled_in_program',
                    target_id=f"program:{project.id}",
                    actor='system',
                )

    return registry


def _issue_receipt_for_donation(donation: Donation, recipient_email: str | None = None):
    """Delegate receipt generation to DonationService.generate_receipt()."""
    svc = DonationService()
    return svc.generate_receipt(
        donation_id=donation.id,
        org_id=donation.organization_id,
        sent_to_email=recipient_email,
    )


@main_bp.route('/api/openapi.yaml', methods=['GET'])
@login_required
@roles_required('admin', 'staff', 'viewer')
def api_openapi_spec():
    return main_api_docs_handlers.api_openapi_spec(_openapi_spec_path())


@main_bp.route('/api/docs', methods=['GET'])
@login_required
@roles_required('admin', 'staff', 'viewer')
def api_docs_index():
    spec_url = url_for('main.api_openapi_spec')
    swagger_url = url_for('main.api_swagger_ui')
    return main_api_docs_handlers.api_docs_index(spec_url=spec_url, swagger_url=swagger_url)


@main_bp.route('/api/swagger', methods=['GET'])
@login_required
@roles_required('admin', 'staff', 'viewer')
def api_swagger_ui():
    spec_url = url_for('main.api_openapi_spec')
    return main_api_docs_handlers.api_swagger_ui(spec_url=spec_url)


@main_bp.route('/api/domain/snapshot', methods=['GET'])
@login_required
@roles_required('admin', 'staff', 'viewer')
def api_domain_snapshot():
    org = _current_org()
    registry = _build_domain_registry_for_org(org)
    return {'ok': True, 'organization': org.name if org else None, 'entities': registry.snapshot()}


@main_bp.route('/api/semantic/context', methods=['GET'])
@login_required
@roles_required('admin', 'staff', 'viewer')
def api_semantic_context():
    task = (request.args.get('task') or '').strip()
    if not task:
        return {'error': 'task query parameter is required.'}, 400

    org = _current_org()
    registry = _build_domain_registry_for_org(org)
    memory = SemanticMemoryLayer()
    org_id = org.id if org else None
    memory.index_registry(registry, organization_id=org_id)
    return {'ok': True, 'context': memory.assemble_context(task=task, limit=6, organization_id=org_id)}


@main_bp.route('/api/ui/profile', methods=['GET', 'PATCH'])
@login_required
@roles_required('admin', 'staff', 'viewer')
def ui_profile_api():
    profile = _normalize_ui_profile(getattr(current_user, "ui_profile_json", None) or _UI_PROFILE_DEFAULT)

    if request.method == 'PATCH':
        payload = request.get_json(silent=True)
        if not isinstance(payload, dict):
            return jsonify({'error': 'JSON object payload is required'}), 400
        profile = _normalize_ui_profile(payload)
        current_user.ui_profile_json = profile
        db.session.commit()

    org = _current_org()
    urgency = _build_operational_urgency(org.id) if org else {'tasks': 0, 'approvals': 0, 'alerts': 0}
    return jsonify({'profile': profile, 'urgency': urgency})


@main_bp.route('/workflows', methods=['GET'])
@login_required
@roles_required('admin', 'staff', 'viewer')
def workflows_page():
    return main_workflow_handlers.workflows_page()


@main_bp.route('/workflows/donation', methods=['POST'])
@login_required
@roles_required('admin', 'staff')
def workflow_donation_route():
    donation_id = request.form.get('donation_id', type=int)
    if not donation_id:
        flash('Donation ID is required.', 'error')
        return redirect(url_for('main.workflows_page'))
    org = _current_org()
    if not org:
        flash('No organization is available.', 'error')
        return redirect(url_for('main.workflows_page'))

    result = main_workflow_handlers.workflow_donation_run(
        donation_id=donation_id,
        actor=getattr(current_user, 'username', 'workflow'),
        org_id=org.id,
        db_path=_sqlite_db_file_path(),
    )
    flash('Workflow completed.' if result.get('ok') else str(result.get('error')), 'success' if result.get('ok') else 'error')
    return redirect(url_for('main.workflows_page'))


@main_bp.route('/workflows/grant', methods=['POST'])
@login_required
@roles_required('admin', 'staff')
def workflow_grant_route():
    grant_name = (request.form.get('grant_name') or '').strip()
    requested_amount = request.form.get('requested_amount', type=float) or 0.0
    if not grant_name:
        flash('Grant name is required.', 'error')
        return redirect(url_for('main.workflows_page'))

    result = main_workflow_handlers.workflow_grant_run(
        grant_name=grant_name,
        requested_amount=requested_amount,
        actor=getattr(current_user, 'username', 'workflow'),
        db_path=_sqlite_db_file_path(),
    )
    flash('Grant workflow completed.' if result.get('ok') else 'Grant workflow failed.', 'success' if result.get('ok') else 'error')
    return redirect(url_for('main.workflows_page'))


@main_bp.route('/workflows/program-impact', methods=['POST'])
@login_required
@roles_required('admin', 'staff')
def workflow_program_route():
    program_name = (request.form.get('program_name') or '').strip()
    beneficiary_count = request.form.get('beneficiary_count', type=int) or 0
    if not program_name:
        flash('Program name is required.', 'error')
        return redirect(url_for('main.workflows_page'))

    result = main_workflow_handlers.workflow_program_run(
        program_name=program_name,
        beneficiary_count=beneficiary_count,
        outcomes=[{'metric_name': 'beneficiary_engagement', 'metric_value': beneficiary_count}],
        actor=getattr(current_user, 'username', 'workflow'),
        db_path=_sqlite_db_file_path(),
    )
    flash('Program workflow completed.' if result.get('ok') else 'Program workflow failed.', 'success' if result.get('ok') else 'error')
    return redirect(url_for('main.workflows_page'))


@main_bp.route('/api/workflows/donation/<int:donation_id>/run', methods=['POST'])
@login_required
@roles_required('admin', 'staff')
def api_workflow_donation_run(donation_id: int):
    org = _current_org()
    if not org:
        return {'ok': False, 'error': 'No organization is available.'}, 400

    result = main_workflow_handlers.workflow_donation_run(
        donation_id=donation_id,
        actor=getattr(current_user, 'username', 'workflow'),
        org_id=org.id,
        db_path=_sqlite_db_file_path(),
    )
    return (result, 200) if result.get('ok') else (result, 404)


@main_bp.route('/api/workflows/grant/run', methods=['POST'])
@login_required
@roles_required('admin', 'staff')
def api_workflow_grant_run():
    payload = request.get_json(silent=True) or {}
    grant_name = str(payload.get('grant_name', '')).strip()
    requested_amount = float(payload.get('requested_amount', 0.0) or 0.0)
    if not grant_name:
        return {'error': 'grant_name is required.'}, 400
    return main_workflow_handlers.workflow_grant_run(
        grant_name=grant_name,
        requested_amount=requested_amount,
        actor=getattr(current_user, 'username', 'workflow'),
        db_path=_sqlite_db_file_path(),
    )


@main_bp.route('/api/workflows/program-impact/run', methods=['POST'])
@login_required
@roles_required('admin', 'staff')
def api_workflow_program_run():
    payload = request.get_json(silent=True) or {}
    program_name = str(payload.get('program_name', '')).strip()
    beneficiary_count = int(payload.get('beneficiary_count', 0) or 0)
    outcomes = payload.get('outcomes') if isinstance(payload.get('outcomes'), list) else []
    if not program_name:
        return {'error': 'program_name is required.'}, 400
    return main_workflow_handlers.workflow_program_run(
        program_name=program_name,
        beneficiary_count=beneficiary_count,
        outcomes=outcomes,
        actor=getattr(current_user, 'username', 'workflow'),
        db_path=_sqlite_db_file_path(),
    )


@main_bp.route('/')
def index():
    """Home/landing page."""
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    return render_template('index.html')


@main_bp.route('/locale/<string:lang>', methods=['POST'])
def set_locale(lang: str):
    normalized = (lang or '').strip().lower()
    if normalized in _SUPPORTED_LOCALES:
        session['lang'] = normalized

    next_path = (request.args.get('next') or request.form.get('next') or request.referrer or url_for('main.index')).strip()
    if not next_path.startswith('/'):
        next_path = url_for('main.index')
    return redirect(next_path)


@main_bp.route('/give', methods=['GET', 'POST'])
def public_give():
    """Public donation page for self-service online giving."""
    org = get_first_active_organization()
    if not org:
        flash('Donation portal is not available yet. Please contact the organization.', 'error')
        return redirect(url_for('main.index'))

    form = PublicDonationForm()
    if request.method == 'GET':
        purpose_prefill = (request.args.get('purpose') or '').strip()
        amount_prefill = (request.args.get('amount') or '').strip()
        if purpose_prefill and not form.purpose.data:
            form.purpose.data = purpose_prefill[:120]
        if amount_prefill and not form.amount.data:
            try:
                parsed_amount = float(amount_prefill)
            except ValueError:
                parsed_amount = 0.0
            if parsed_amount > 0:
                form.amount.data = round(parsed_amount, 2)

    if form.validate_on_submit():
        donor_email = (form.donor_email.data or '').strip() or None

        _donor_svc = DonorService()
        donor, _created = _donor_svc.find_or_create_by_email(
            org.id,
            donor_email,
            form.donor_name.data.strip(),
            phone=(form.donor_phone.data or '').strip() or None,
            donor_type='individual',
            notes='Created from public donation portal.',
        )

        _donation_svc = DonationService()
        donation = _donation_svc.create_donation(
            org_id=org.id,
            donor_name=donor.name,
            amount=form.amount.data,
            currency=form.currency.data,
            donor_email=donor.email,
            donor_phone=donor.phone,
            donor_id=donor.id,
            payment_method=form.payment_method.data,
            purpose=form.purpose.data or 'General Fund',
            notes='Public portal donation',
            status='received',
        )

        if form.make_recurring.data:
            if not donor.email:
                flash('Recurring donations require an email to process retries and receipts.', 'error')
                return render_template('public_donation_form.html', form=form, active_page='give')
            _donation_svc.create_recurring_plan(
                org.id,
                donor.id,
                amount=form.amount.data,
                currency=form.currency.data,
                payment_method=form.payment_method.data,
                purpose=form.purpose.data or 'General Fund',
                frequency=form.recurring_frequency.data or 'monthly',
                next_charge_date=_next_charge_date(date.today(), form.recurring_frequency.data or 'monthly'),
            )

        # Advance to 'processed' so generate_receipt() is satisfied, then issue.
        _donation_svc.update_status(donation.id, org.id, 'processed', actor_id=None)
        _issue_receipt_for_donation(donation, recipient_email=donor.email)

        flash('Thank you for your donation. Your receipt has been generated.', 'success')
        return redirect(url_for('main.public_give'))

    return render_template('public_donation_form.html', form=form, active_page='give')


@main_bp.route('/public/donate', methods=['GET'])
def public_donate_page():
    org = get_first_active_organization()
    campaign = None
    campaigns: list[Campaign] = []
    funds: list[Fund] = []
    is_embed = request.args.get('embed', '0') == '1'
    if org is not None:
        campaigns = list(
            db.session.scalars(
                select(Campaign)
                .where(Campaign.organization_id == org.id, Campaign.status == 'active')
                .order_by(Campaign.name.asc(), Campaign.id.asc())
            )
        )
        campaign = campaigns[0] if campaigns else None
        funds = FundService().list_all_funds(org.id, active_only=True)
    template_name = 'public/donate_embed.html' if is_embed else 'public/donate.html'
    return render_template(
        template_name,
        campaign=campaign,
        campaigns=campaigns,
        funds=funds,
        is_embed=is_embed,
    )


@main_bp.route('/public/donate/embed.js', methods=['GET'])
def public_donate_embed_script():
    src = '/public/donate?embed=1'
    safe_src = json.dumps(src)
    safe_title = json.dumps('Donate Securely')
    script = f"""
(function() {{
  var script = document.currentScript;
  if (!script) return;
  var targetId = script.getAttribute('data-target');
  var target = targetId ? document.getElementById(targetId) : script.parentNode;
  if (!target) return;
  var iframe = document.createElement('iframe');
  iframe.src = {safe_src};
  iframe.title = {safe_title};
  iframe.width = '100%';
  iframe.height = script.getAttribute('data-height') || '860';
  iframe.style.border = '0';
  iframe.style.maxWidth = '100%';
  iframe.style.borderRadius = '12px';
  iframe.loading = 'lazy';
  target.appendChild(iframe);
}})();
""".strip()
    return Response(script, mimetype='application/javascript')


@main_bp.route('/api/events/<int:event_id>/validate-code', methods=['POST'])
def validate_event_discount_code(event_id: int):
    payload = request.get_json(silent=True) or {}
    code = str(payload.get('code') or '').strip()
    amount = float(payload.get('amount') or 0.0)

    if not code:
        return {'valid': False, 'error': 'code is required'}, 400
    if amount <= 0:
        return {'valid': False, 'error': 'amount must be greater than zero'}, 400

    discount = _resolve_event_discount_code(event_id, code)
    if discount is None:
        return {'valid': False, 'error': 'Invalid or expired discount code'}, 404

    discount_value = _discount_amount(amount, discount)
    final_amount = round(max(0.0, amount - discount_value), 2)
    return {
        'valid': True,
        'code': discount.code,
        'discount_type': discount.discount_type,
        'discount_value': float(discount.discount_value or 0.0),
        'discount_amount': discount_value,
        'final_amount': final_amount,
    }, 200


@main_bp.route('/public/donate/create-checkout', methods=['POST'])
def public_donate_create_checkout():
    org = get_first_active_organization()
    if not org:
        flash('Donation portal is unavailable.', 'error')
        return redirect(url_for('main.index'))

    amount = float(request.form.get('amount', '0') or 0)
    donor_email = (request.form.get('email') or '').strip().lower()
    donor_name = (request.form.get('donor_name') or '').strip()
    purpose = (request.form.get('purpose') or '').strip()
    raw_campaign_id = (request.form.get('campaign_id') or '').strip()
    raw_fund_id = (request.form.get('fund_id') or '').strip()
    raw_event_id = (request.form.get('event_id') or '').strip()
    discount_code = (request.form.get('discount_code') or '').strip()
    tribute_type_raw = (request.form.get('tribute_type') or '').strip().lower()
    tribute_honoree_name = (request.form.get('tribute_honoree_name') or '').strip()
    tribute_honoree_contact = (request.form.get('tribute_honoree_contact') or '').strip()
    campaign_id = int(raw_campaign_id) if raw_campaign_id.isdigit() else None
    fund_id = int(raw_fund_id) if raw_fund_id.isdigit() else None
    event_id = int(raw_event_id) if raw_event_id.isdigit() else None
    tribute_type = tribute_type_raw if tribute_type_raw in {'in_honor_of', 'in_memory_of'} else None

    if amount <= 0:
        flash('Amount must be greater than zero.', 'error')
        return redirect(url_for('main.public_donate_page'))
    if not donor_email:
        flash('Email is required.', 'error')
        return redirect(url_for('main.public_donate_page'))

    if campaign_id is not None:
        campaign = db.session.scalars(
            select(Campaign).where(Campaign.id == campaign_id, Campaign.organization_id == org.id).limit(1)
        ).first()
        if campaign is None:
            flash('Selected campaign is not available.', 'error')
            return redirect(url_for('main.public_donate_page'))

    if fund_id is not None and fund_id > 0:
        fund = db.session.scalars(
            select(Fund).where(Fund.id == fund_id, Fund.organization_id == org.id, Fund.is_active == True).limit(1)
        ).first()
        if fund is None:
            flash('Selected fund is not available.', 'error')
            return redirect(url_for('main.public_donate_page'))
    else:
        fund_id = None

    if tribute_type and not tribute_honoree_name:
        flash('Honoree name is required for tribute gifts.', 'error')
        return redirect(url_for('main.public_donate_page'))

    attempts_in_last_hour = _public_donation_attempts_last_hour(org_id=org.id, donor_email=donor_email)
    if attempts_in_last_hour >= _PUBLIC_DONATION_MAX_ATTEMPTS_PER_HOUR:
        flash('Too many donation attempts for this email. Please wait and try again later.', 'error')
        return redirect(url_for('main.public_donate_page'))

    discount = None
    discount_amount = 0.0
    final_amount = amount
    if event_id is not None and discount_code:
        discount = _resolve_event_discount_code(event_id, discount_code)
        if discount is None:
            flash('Invalid or expired discount code.', 'error')
            return redirect(url_for('main.public_donate_page'))
        discount_amount = _discount_amount(amount, discount)
        final_amount = round(max(0.5, amount - discount_amount), 2)

    donor_service = DonorService()
    donor, _created = donor_service.find_or_create_by_email(
        org.id,
        donor_email,
        donor_name or donor_email.split('@')[0] or 'Donor',
        donor_type='individual',
        notes='Created from public Stripe donate flow.',
    )

    donation = DonationService().create_donation(
        org_id=org.id,
        donor_name=donor.name,
        amount=final_amount,
        currency='USD',
        donor_email=donor.email,
        donor_id=donor.id,
        campaign_id=campaign_id,
        fund_id=fund_id,
        payment_method='stripe',
        purpose=purpose or 'Public Stripe donation',
        tribute_type=tribute_type,
        tribute_honoree_name=tribute_honoree_name or None,
        tribute_honoree_contact=tribute_honoree_contact or None,
        notes=(
            f'Pending Stripe checkout. Discount code {discount.code} applied: -${discount_amount:.2f}'
            if discount is not None
            else 'Pending Stripe checkout'
        ),
        status='pending',
    )

    campaign_slug = 'general'
    if campaign_id:
        campaign = db.session.scalars(
            select(Campaign).where(Campaign.id == campaign_id, Campaign.organization_id == org.id).limit(1)
        ).first()
        if campaign is not None:
            campaign_slug = campaign.slug

    try:
        checkout = PaymentService().create_checkout_session(
            org_id=org.id,
            donor_id=donor.id,
            campaign_id=campaign_id,
            event_id=event_id,
            donation_id=donation.id,
            amount_cents=int(round(final_amount * 100)),
            currency='USD',
            campaign_name=f'Donation to {campaign_slug}',
            success_url=url_for('main.public_donate_success', donation_id=donation.id, _external=True),
            cancel_url=url_for('main.public_donate_cancel', donation_id=donation.id, _external=True),
            donor_email=donor.email,
            donor_name=donor.name,
        )
    except StripeNotConfigured:
        flash('Online card payments are not configured yet.', 'error')
        return redirect(url_for('main.public_donate_page'))

    return redirect(checkout['checkout_url'])


@main_bp.route('/public/donate/success', methods=['GET'])
def public_donate_success():
    donation_id = request.args.get('donation_id', '')
    return render_template('index.html', success_message=f'Thank you. Donation #{donation_id} is being finalized.')


@main_bp.route('/public/donate/cancel', methods=['GET'])
def public_donate_cancel():
    donation_id = request.args.get('donation_id', '')
    return render_template('index.html', error_message=f'Donation #{donation_id} was canceled.')


@main_bp.route('/setup', methods=['GET', 'POST'])
@login_required
@roles_required('admin')
def org_setup():
    """First-time organization setup wizard."""
    from ngo_homesuite.models.core import Organization, User, db
    import re

    org = _current_org()
    if org:
        flash('Organization already configured. Use Settings to make changes.', 'info')
        return redirect(url_for('main.org_settings'))

    error = None
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip() or None
        phone = request.form.get('phone', '').strip() or None
        website = request.form.get('website', '').strip() or None
        country = request.form.get('country', '').strip() or None
        city = request.form.get('city', '').strip() or None
        mission = request.form.get('mission', '').strip() or None

        if not name:
            error = 'Organization name is required.'
        else:
            slug = re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-') or 'org'
            # ensure unique slug
            existing = db.session.query(Organization).filter_by(slug=slug).first()
            if existing:
                slug = f"{slug}-{existing.id + 1}"

            new_org = Organization(
                name=name, slug=slug, email=email, phone=phone,
                website=website, country=country, city=city,
                mission=mission, is_active=True,
            )
            db.session.add(new_org)
            db.session.flush()  # get new_org.id

            # Assign org to current user
            user = db.session.get(User, current_user.id)
            user.organization_id = new_org.id
            db.session.commit()
            flash(f'Organization "{name}" created successfully.', 'success')
            return redirect(url_for('main.dashboard'))

    return render_template('setup.html', error=error, active_page=None)


@main_bp.route('/settings', methods=['GET', 'POST'])
@login_required
@roles_required('admin')
def org_settings():
    """Organization settings page."""
    from ngo_homesuite.models.core import Organization, db

    org = _current_org()
    error = None
    success = False

    if request.method == 'POST' and org:
        import re
        name = request.form.get('name', '').strip()
        if not name:
            error = 'Organization name is required.'
        else:
            org.name = name
            org.email = request.form.get('email', '').strip() or None
            org.phone = request.form.get('phone', '').strip() or None
            org.website = request.form.get('website', '').strip() or None
            org.country = request.form.get('country', '').strip() or None
            org.city = request.form.get('city', '').strip() or None
            org.mission = request.form.get('mission', '').strip() or None
            db.session.commit()
            flash('Settings saved.', 'success')
            success = True

    return render_template('settings.html', org=org, error=error, success=success, active_page='settings')


@main_bp.route('/dashboard')
@login_required
def dashboard():
    """User dashboard with summary cards."""

    selected_period = str(request.args.get('period', '30d') or '30d').strip().lower()
    if selected_period not in {'30d', '90d', 'ytd', 'custom'}:
        selected_period = '30d'

    start_date_raw = str(request.args.get('start_date', '') or '').strip()
    end_date_raw = str(request.args.get('end_date', '') or '').strip()
    custom_start_dt = None
    custom_end_dt = None
    if selected_period == 'custom' and start_date_raw and end_date_raw:
        try:
            custom_start_dt = datetime.strptime(start_date_raw, '%Y-%m-%d')
            custom_end_dt = datetime.strptime(end_date_raw, '%Y-%m-%d')
            if custom_end_dt < custom_start_dt:
                custom_start_dt = None
                custom_end_dt = None
                selected_period = '30d'
        except ValueError:
            custom_start_dt = None
            custom_end_dt = None
            selected_period = '30d'

    org = _current_org()
    if org:
        summary = ReportingService().organization_dashboard_summary(
            org.id,
            recent_donations_limit=5,
            period=selected_period,
            start_date=custom_start_dt,
            end_date=custom_end_dt,
        )
        stats = {
            'organization': org,
            **summary,
        }
    else:
        stats = {
            'organization': None,
            'beneficiary_count': 0,
            'project_count': 0,
            'donor_count': 0,
            'total_donations': 0,
            'total_budget': 0,
            'total_expenses': 0,
            'net_cashflow': 0,
            'total_funds': 0,
            'donation_transaction_count': 0,
            'avg_gift_amount': 0,
            'recurring_active_plans': 0,
            'recurring_monthly_commitment': 0,
            'recent_donations': [],
            'trend_30d': {
                'donations': 0,
                'expenses': 0,
                'net': 0,
                'donations_delta_pct': 0,
                'expenses_delta_pct': 0,
            },
            'trend_90d': {
                'donations': 0,
                'expenses': 0,
                'net': 0,
                'donations_delta_pct': 0,
                'expenses_delta_pct': 0,
            },
            'trend_ytd': {
                'donations': 0,
                'expenses': 0,
                'net': 0,
                'donations_delta_pct': 0,
                'expenses_delta_pct': 0,
            },
            'selected_period': selected_period,
            'selected_period_label': {
                '30d': 'Last 30 Days',
                '90d': 'Last 90 Days',
                'ytd': 'Year to Date',
                'custom': 'Custom Range',
            }.get(selected_period, 'Last 30 Days'),
            'period_focus': {
                'donations': 0,
                'expenses': 0,
                'net': 0,
                'donations_delta_pct': 0,
                'expenses_delta_pct': 0,
                'net_delta_pct': 0,
            },
            'period_focus_series': {
                'donations': [0, 0, 0, 0, 0, 0],
                'expenses': [0, 0, 0, 0, 0, 0],
            },
            'period_comparison': {
                'label': '90-Day Run Rate (30D Equivalent)',
                'donations': 0,
                'expenses': 0,
                'net': 0,
                'donations_delta_pct': 0,
                'expenses_delta_pct': 0,
                'net_delta_pct': 0,
            },
            'period_comparison_series': {
                'donations': [0, 0, 0, 0, 0, 0],
                'expenses': [0, 0, 0, 0, 0, 0],
            },
            'custom_range': {
                'start_date': start_date_raw or None,
                'end_date': end_date_raw or None,
            },
            'monthly_overview': {
                'labels': [],
                'donations': [],
                'expenses': [],
                'net': [],
                'max_value': 1,
            },
            'category_metrics': {
                'fundraising': {
                    'total_donations': 0,
                    'transactions': 0,
                    'avg_gift': 0,
                    'recurring_monthly_commitment': 0,
                    'score': 0,
                },
                'financial_health': {
                    'net_cashflow': 0,
                    'expense_ratio_pct': 0,
                    'net_margin_pct': 0,
                    'score': 0,
                },
                'program_delivery': {
                    'active_projects': 0,
                    'active_beneficiaries': 0,
                    'beneficiaries_per_project': 0,
                    'score': 0,
                },
                'donor_engagement': {
                    'total_donors': 0,
                    'active_recurring_plans': 0,
                    'recurring_penetration_pct': 0,
                    'score': 0,
                },
                'operations': {
                    'missing_donation_dates': 0,
                    'days_since_last_donation': 0,
                    'data_completeness_pct': 0,
                    'score': 0,
                },
            },
            'goal_progress': {
                'fundraising_goal': 0,
                'fundraising_progress': 0,
                'expense_cap': 0,
                'expense_progress': 0,
            },
            'donor_lifecycle': {
                'window_days': 30,
                'current_active_donors': 0,
                'new_donors': 0,
                'returning_donors': 0,
                'retained_donors': 0,
                'lapsed_donors': 0,
                'reactivated_donors': 0,
                'retention_pct': 0,
                'lapse_pct': 0,
            },
            'campaign_attribution': {
                'top_campaigns': [],
                'attributed_amount': 0,
                'unattributed_amount': 0,
                'unattributed_donations': 0,
                'coverage_pct': 0,
            },
            'budget_variance': {
                'window_label': 'Last 30 Days',
                'project_budget_total': 0,
                'project_spent_total': 0,
                'project_variance_total': 0,
                'over_budget_projects': 0,
                'top_projects': [],
            },
            'forecast': {
                'month_label': datetime.utcnow().strftime('%B %Y'),
                'days_elapsed': 0,
                'days_in_month': 0,
                'mtd_donations': 0,
                'mtd_expenses': 0,
                'donation_daily_run_rate': 0,
                'expense_daily_run_rate': 0,
                'projected_month_donations': 0,
                'projected_month_expenses': 0,
                'projected_month_net': 0,
            },
            'donor_cohorts': {
                'labels': [],
                'new': [],
                'returning': [],
                'retained': [],
            },
            'alerts': [],
            'data_freshness': {
                'generated_at': None,
                'latest_donation_date': None,
            },
            'data_quality': {
                'missing_donation_dates': 0,
            },
        }

    period_params: dict[str, str] = {'period': selected_period}
    if selected_period == 'custom' and stats.get('custom_range', {}).get('start_date') and stats.get('custom_range', {}).get('end_date'):
        period_params['start_date'] = str(stats['custom_range']['start_date'])
        period_params['end_date'] = str(stats['custom_range']['end_date'])
    export_url = url_for('main.dashboard_export', **period_params)

    metric_map = {
        'donor_count': {
            'label': 'Donors',
            'value': str(stats['donor_count']),
            'link': url_for('main.donors_list'),
        },
        'beneficiary_count': {
            'label': 'Active Beneficiaries',
            'value': str(stats['beneficiary_count']),
            'link': url_for('main.beneficiaries_list'),
        },
        'project_count': {
            'label': 'Active Projects',
            'value': str(stats['project_count']),
            'link': url_for('main.projects_dashboard', **period_params),
        },
        'total_donations': {
            'label': 'Total Donations',
            'value': f"${stats['total_donations']:.2f}",
            'link': url_for('main.donations_list', **period_params),
        },
        'total_budget': {
            'label': 'Project Budget',
            'value': f"${stats['total_budget']:.2f}",
            'link': url_for('main.projects_dashboard', **period_params),
        },
        'total_expenses': {
            'label': 'Total Expenses',
            'value': f"${stats['total_expenses']:.2f}",
            'link': url_for('main.expenses_list', **period_params),
        },
        'total_funds': {
            'label': 'Active Funds',
            'value': str(stats['total_funds']),
            'link': url_for('main.funds_list'),
        },
    }
    role = str(getattr(current_user, 'role', 'viewer') or 'viewer')
    role_card_order = {
        'admin': ['total_donations', 'total_expenses', 'total_budget', 'total_funds', 'project_count', 'donor_count', 'beneficiary_count'],
        'staff': ['project_count', 'beneficiary_count', 'donor_count', 'total_donations', 'total_expenses', 'total_funds', 'total_budget'],
        'volunteer': ['beneficiary_count', 'project_count', 'donor_count', 'total_donations', 'total_funds', 'total_expenses', 'total_budget'],
        'viewer': ['donor_count', 'beneficiary_count', 'project_count', 'total_donations', 'total_budget', 'total_expenses', 'total_funds'],
    }
    ordered_metric_cards = [metric_map[key] for key in role_card_order.get(role, role_card_order['viewer']) if key in metric_map]
    
    ai_context = {
        'active_page': 'dashboard',
        'organization': org.name if org else None,
        'donor_count': stats['donor_count'],
        'total_donations': stats['total_donations'],
        'total_expenses': stats['total_expenses'],
        'project_count': stats['project_count'],
        'total_funds': stats['total_funds'],
    }
    return render_template(
        'dashboard.html',
        stats=stats,
        active_page='dashboard',
        ai_context=ai_context,
        metric_cards=ordered_metric_cards,
        dashboard_role=role,
        selected_period=selected_period,
        selected_start_date=stats.get('custom_range', {}).get('start_date') or start_date_raw,
        selected_end_date=stats.get('custom_range', {}).get('end_date') or end_date_raw,
        export_url=export_url,
        period_params=period_params,
    )


@main_bp.route('/dashboard/export')
@login_required
@require_step_up_auth
def dashboard_export() -> Response:
    """Export the current dashboard snapshot as JSON."""
    selected_period = str(request.args.get('period', '30d') or '30d').strip().lower()
    if selected_period not in {'30d', '90d', 'ytd', 'custom'}:
        selected_period = '30d'

    start_date_raw = str(request.args.get('start_date', '') or '').strip()
    end_date_raw = str(request.args.get('end_date', '') or '').strip()
    custom_start_dt = None
    custom_end_dt = None
    if selected_period == 'custom' and start_date_raw and end_date_raw:
        try:
            custom_start_dt = datetime.strptime(start_date_raw, '%Y-%m-%d')
            custom_end_dt = datetime.strptime(end_date_raw, '%Y-%m-%d')
            if custom_end_dt < custom_start_dt:
                custom_start_dt = None
                custom_end_dt = None
                selected_period = '30d'
        except ValueError:
            custom_start_dt = None
            custom_end_dt = None
            selected_period = '30d'

    org = _current_org()
    if not org:
        return Response(
            json.dumps({'error': 'No organization found for current user'}),
            status=400,
            mimetype='application/json',
        )

    summary = ReportingService().organization_dashboard_summary(
        org.id,
        recent_donations_limit=5,
        period=selected_period,
        start_date=custom_start_dt,
        end_date=custom_end_dt,
    )
    payload = {
        'organization': {
            'id': org.id,
            'name': org.name,
        },
        'filters': {
            'period': selected_period,
            'start_date': start_date_raw or None,
            'end_date': end_date_raw or None,
            'generated_at': datetime.utcnow().isoformat(timespec='seconds'),
        },
        'summary': summary,
    }
    filename = f"dashboard-snapshot-{selected_period}-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.json"
    return Response(
        json.dumps(payload, default=str, indent=2),
        status=200,
        mimetype='application/json',
        headers={
            'Content-Disposition': f'attachment; filename="{filename}"',
        },
    )


@main_bp.route('/activity')
@login_required
def activity_feed():
    """Organization-wide activity feed dashboard."""
    org = _current_org()
    if not org:
        flash('No organization found for your account.', 'error')
        return redirect(url_for('main.dashboard'))

    return render_template(
        'activity_feed.html',
        active_page='activity',
        organization_name=org.name,
    )


@main_bp.route('/tony-scoring')
@login_required
@roles_required('admin', 'staff')
def tony_scoring():
    """TONY advanced grant scoring dashboard."""
    org = _current_org()
    if not org:
        flash('No organization found for your account.', 'error')
        return redirect(url_for('main.dashboard'))

    if 'tony.tony_home' in current_app.view_functions:
        return redirect(url_for('tony.tony_home'))

    return render_template(
        'tony_scoring.html',
        active_page='tony_scoring',
        organization_name=org.name,
    )


@main_bp.route('/tasks/board')
@login_required
def task_board():
    """Operational task board powered by v2 task + reminder APIs."""
    org = _current_org()
    if not org:
        flash('No organization found for your account.', 'error')
        return redirect(url_for('main.dashboard'))

    ai_context = {
        'active_page': 'task_board',
        'organization': org.name,
    }
    return render_template(
        'task_board.html',
        active_page='task_board',
        organization_name=org.name,
        ai_context=ai_context,
    )


@main_bp.route('/mobile/intake', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'staff', 'volunteer')
def mobile_intake():
    org = _current_org()
    if not org:
        flash('No organization found for your account.', 'error')
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        action = (request.form.get('action') or '').strip().lower()
        if action == 'beneficiary':
            first_name = (request.form.get('first_name') or '').strip()
            last_name = (request.form.get('last_name') or '').strip()
            if not first_name or not last_name:
                flash('Beneficiary first and last name are required.', 'error')
            else:
                create_beneficiary(
                    org.id,
                    first_name,
                    last_name,
                    phone=(request.form.get('phone') or '').strip() or None,
                    city=(request.form.get('city') or '').strip() or None,
                    program=(request.form.get('program') or '').strip() or None,
                    status=(request.form.get('status') or 'active').strip() or 'active',
                    notes=(request.form.get('notes') or '').strip() or None,
                )
                flash('Beneficiary quick intake captured.', 'success')
                return redirect(url_for('main.mobile_intake'))
        elif action == 'volunteer':
            name = (request.form.get('volunteer_name') or '').strip()
            if not name:
                flash('Volunteer name is required.', 'error')
            else:
                create_volunteer(
                    org.id,
                    name,
                    email=(request.form.get('volunteer_email') or '').strip() or None,
                    phone=(request.form.get('volunteer_phone') or '').strip() or None,
                    status=(request.form.get('volunteer_status') or 'active').strip() or 'active',
                )
                flash('Volunteer quick registration captured.', 'success')
                return redirect(url_for('main.mobile_intake'))
        else:
            flash('Unsupported intake action.', 'error')

    recent_beneficiaries = list_beneficiaries(org.id)[:8]
    recent_volunteers = list_recent_volunteers(org.id, limit=8)

    ai_context = {
        'active_page': 'mobile_intake',
        'organization': org.name,
        'recent_beneficiaries': len(recent_beneficiaries),
        'recent_volunteers': len(recent_volunteers),
    }
    return render_template(
        'mobile_intake.html',
        active_page='mobile_intake',
        beneficiaries=recent_beneficiaries,
        volunteers=recent_volunteers,
        ai_context=ai_context,
    )


@main_bp.route('/p2p/manage', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'staff')
def p2p_manage():
    from ngo_homesuite.services.p2p_service import close_page, create_page, get_progress, list_pages, publish_page, update_page

    org = _current_org()
    if not org:
        flash('No organization found for your account.', 'error')
        return redirect(url_for('main.dashboard'))

    form = P2PPageForm()
    donors = DonorService().list_all_donors(org.id)
    form.donor_id.choices = [(int(d.id), d.name) for d in donors]
    donor_name_by_id = {int(d.id): d.name for d in donors}

    status_filter = (request.args.get('status') or '').strip().lower() or None
    owner_filter = request.args.get('owner_id', type=int)
    campaign_filter = (request.args.get('campaign') or '').strip().lower()
    query = (request.args.get('q') or '').strip().lower()
    sort_by = (request.args.get('sort_by') or 'created').strip().lower()
    sort_dir = (request.args.get('sort_dir') or 'desc').strip().lower()
    if sort_by not in {'created', 'title', 'raised', 'progress', 'supporters', 'recent30'}:
        sort_by = 'created'
    if sort_dir not in {'asc', 'desc'}:
        sort_dir = 'desc'

    pages = list_pages(org.id, status=status_filter)

    if request.method == 'POST':
        action = (request.form.get('action') or '').strip().lower()
        page_id = request.form.get('page_id', type=int)

        if action in {'publish', 'close'}:
            if not page_id:
                flash('A valid fundraiser page id is required.', 'error')
                return redirect(url_for('main.p2p_manage'))

            try:
                if action == 'publish':
                    publish_page(page_id, org.id)
                    flash('Fundraiser published.', 'success')
                else:
                    close_page(page_id, org.id)
                    flash('Fundraiser closed.', 'success')
            except NotFound:
                flash('Fundraiser page not found for this organization.', 'error')
            except ValueError as exc:
                flash(str(exc), 'error')
            return redirect(url_for('main.p2p_manage'))

        if action == 'bulk_status':
            selected_ids = [
                int(raw_id)
                for raw_id in request.form.getlist('selected_page_ids')
                if str(raw_id).isdigit()
            ]
            bulk_target = (request.form.get('bulk_target_status') or '').strip().lower()
            if not selected_ids:
                flash('Select at least one fundraiser for a bulk action.', 'error')
                return redirect(url_for('main.p2p_manage'))
            if bulk_target not in {'active', 'closed'}:
                flash('Unsupported bulk status action.', 'error')
                return redirect(url_for('main.p2p_manage'))

            processed = 0
            for selected_id in selected_ids:
                try:
                    if bulk_target == 'active':
                        publish_page(selected_id, org.id)
                    else:
                        close_page(selected_id, org.id)
                    processed += 1
                except Exception:
                    continue
            flash(f'Bulk action applied to {processed} fundraiser(s).', 'success')
            return redirect(url_for('main.p2p_manage'))

        if action == 'update':
            if not page_id:
                flash('A valid fundraiser page id is required.', 'error')
                return redirect(url_for('main.p2p_manage'))
            challenge_end_date = None
            raw_challenge_end = (request.form.get('challenge_end_date') or '').strip()
            if raw_challenge_end:
                try:
                    challenge_end_date = date.fromisoformat(raw_challenge_end)
                except ValueError:
                    flash('Challenge end date must be in YYYY-MM-DD format.', 'error')
                    return redirect(url_for('main.p2p_manage'))
            try:
                update_page(
                    page_id,
                    org.id,
                    title=(request.form.get('title') or '').strip(),
                    story=(request.form.get('story') or '').strip() or None,
                    goal_amount=float(request.form.get('goal_amount') or 0.0),
                    campaign_slug=(request.form.get('campaign_slug') or '').strip() or None,
                    match_ratio=float(request.form.get('match_ratio') or 0.0),
                    match_cap_amount=float(request.form.get('match_cap_amount') or 0.0),
                    challenge_goal_amount=float(request.form.get('challenge_goal_amount') or 0.0),
                    challenge_end_date=challenge_end_date,
                    automation_contact_email=(request.form.get('automation_contact_email') or '').strip() or None,
                )
                flash('Fundraiser updated.', 'success')
            except NotFound:
                flash('Fundraiser page not found for this organization.', 'error')
            except ValueError as exc:
                flash(str(exc), 'error')
            return redirect(url_for('main.p2p_manage'))

        if action == 'reassign_owner':
            if not page_id:
                flash('A valid fundraiser page id is required.', 'error')
                return redirect(url_for('main.p2p_manage'))
            new_owner_id = request.form.get('new_owner_id', type=int)
            if not new_owner_id:
                flash('Select a valid owner to reassign.', 'error')
                return redirect(url_for('main.p2p_manage'))
            try:
                update_page(page_id, org.id, donor_id=int(new_owner_id))
                flash('Fundraiser owner reassigned.', 'success')
            except NotFound:
                flash('Fundraiser page not found for this organization.', 'error')
            except ValueError as exc:
                flash(str(exc), 'error')
            return redirect(url_for('main.p2p_manage'))

        if action == 'send_nudge':
            if not page_id:
                flash('A valid fundraiser page id is required.', 'error')
                return redirect(url_for('main.p2p_manage'))
            page = next((p for p in pages if int(p.id) == int(page_id)), None)
            if page is None:
                flash('Fundraiser page not found for this organization.', 'error')
                return redirect(url_for('main.p2p_manage'))
            contact_email = str(getattr(page, 'automation_contact_email', '') or '').strip()
            owner_name = donor_name_by_id.get(int(page.donor_id), 'Fundraiser Owner')
            if contact_email:
                flash(f'Readiness nudge prepared for {owner_name} at {contact_email}.', 'success')
            else:
                flash('No automation contact email is configured for this page.', 'error')
            return redirect(url_for('main.p2p_manage'))

        if not donors:
            flash('Create at least one donor before creating a fundraiser page.', 'error')
            return render_template('p2p_manage.html', form=form, pages=pages, active_page='p2p')

        if form.validate_on_submit():
            try:
                create_page(
                    organization_id=org.id,
                    donor_id=int(form.donor_id.data),
                    title=(form.title.data or '').strip(),
                    goal_amount=float(form.goal_amount.data or 0.0),
                    story=(form.story.data or '').strip() or None,
                    campaign_slug=(form.campaign_slug.data or '').strip() or None,
                    match_ratio=float(form.match_ratio.data or 0.0),
                    match_cap_amount=float(form.match_cap_amount.data or 0.0),
                    challenge_goal_amount=float(form.challenge_goal_amount.data or 0.0),
                    challenge_end_date=form.challenge_end_date.data,
                    automation_contact_email=(form.automation_contact_email.data or '').strip() or None,
                )
                flash('Fundraiser page created.', 'success')
                return redirect(url_for('main.p2p_manage'))
            except ValueError:
                flash('Invalid fundraiser data.', 'error')
        else:
            flash('Please fix the highlighted form issues.', 'error')

    if query:
        pages = [
            p for p in pages
            if query in str(getattr(p, 'title', '') or '').lower()
            or query in str(getattr(p, 'public_slug', '') or '').lower()
            or query in str(getattr(p, 'story', '') or '').lower()
            or query in str(getattr(getattr(p, 'owner', None), 'name', '') or '').lower()
        ]
    if owner_filter:
        pages = [p for p in pages if int(getattr(p, 'donor_id', 0) or 0) == int(owner_filter)]
    if campaign_filter:
        pages = [p for p in pages if campaign_filter in str(getattr(p, 'campaign_slug', '') or '').lower()]

    page_stats: dict[int, dict[str, object]] = {}
    total_raised = 0.0
    pages_at_goal = 0
    total_raised_30d = 0.0
    campaign_rollups: dict[str, dict[str, object]] = {}
    owner_rollups: dict[int, dict[str, object]] = {}
    automation_queue: list[dict[str, object]] = []
    now_utc_naive = datetime.now(timezone.utc).replace(tzinfo=None)

    for page in pages:
        progress = get_progress(int(page.id), int(org.id))
        raised = float(progress.get('total_raised', 0.0) or 0.0)
        pct = float(progress.get('pct_of_goal', 0.0) or 0.0)
        supporters = int(progress.get('donor_count', 0) or 0)
        linked_donations = list(getattr(page, 'donations', []) or [])

        raised_30d = 0.0
        raised_7d = 0.0
        last_donation_at = None
        for donation in linked_donations:
            amount = float(getattr(donation, 'amount', 0.0) or 0.0)
            donation_date = getattr(donation, 'donation_date', None)
            if isinstance(donation_date, datetime):
                dt = donation_date
                if dt.tzinfo is not None:
                    dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
            else:
                continue
            if last_donation_at is None or dt > last_donation_at:
                last_donation_at = dt
            if dt >= now_utc_naive - timedelta(days=30):
                raised_30d += amount
            if dt >= now_utc_naive - timedelta(days=7):
                raised_7d += amount

        total_raised_30d += raised_30d
        match_ratio = float(getattr(page, 'match_ratio', 0.0) or 0.0)
        match_cap = float(getattr(page, 'match_cap_amount', 0.0) or 0.0)
        challenge_goal = float(getattr(page, 'challenge_goal_amount', 0.0) or 0.0)
        matched_value = min(raised * match_ratio, match_cap) if match_ratio > 0 and match_cap > 0 else 0.0
        challenge_pct = round((raised / challenge_goal) * 100, 1) if challenge_goal > 0 else 0.0

        if pct >= 100.0:
            pages_at_goal += 1
        total_raised += raised
        page_stats[int(page.id)] = {
            'raised': round(raised, 2),
            'pct_of_goal': round(pct, 1),
            'supporters': supporters,
            'raised_30d': round(raised_30d, 2),
            'raised_7d': round(raised_7d, 2),
            'matched_value': round(matched_value, 2),
            'challenge_pct': challenge_pct,
            'last_donation_at': last_donation_at,
        }

        campaign_key = str(getattr(page, 'campaign_slug', '') or '').strip() or 'unmapped'
        rollup = campaign_rollups.setdefault(
            campaign_key,
            {'campaign': campaign_key, 'pages': 0, 'raised': 0.0, 'supporters': 0, 'active': 0},
        )
        rollup['pages'] = int(rollup['pages']) + 1
        rollup['raised'] = float(rollup['raised']) + raised
        rollup['supporters'] = int(rollup['supporters']) + supporters
        if str(getattr(page, 'status', '') or '') == 'active':
            rollup['active'] = int(rollup['active']) + 1

        owner_id = int(getattr(page, 'donor_id', 0) or 0)
        owner_rollup = owner_rollups.setdefault(
            owner_id,
            {'owner_id': owner_id, 'owner_name': donor_name_by_id.get(owner_id, 'Unknown'), 'pages': 0, 'raised': 0.0},
        )
        owner_rollup['pages'] = int(owner_rollup['pages']) + 1
        owner_rollup['raised'] = float(owner_rollup['raised']) + raised

        inactive_days = None
        if isinstance(last_donation_at, datetime):
            inactive_days = max((now_utc_naive - last_donation_at).days, 0)
        if str(getattr(page, 'status', '') or '') == 'active' and (
            (inactive_days is not None and inactive_days >= 14)
            or (pct >= 80 and pct < 100)
        ):
            automation_queue.append(
                {
                    'page_id': int(page.id),
                    'title': str(getattr(page, 'title', '') or ''),
                    'inactive_days': inactive_days,
                    'pct': round(pct, 1),
                    'owner_name': donor_name_by_id.get(owner_id, 'Owner'),
                }
            )

    def _sort_key(page):
        stats = page_stats.get(int(page.id), {})
        if sort_by == 'title':
            return str(getattr(page, 'title', '') or '').lower()
        if sort_by == 'raised':
            return float(stats.get('raised', 0.0) or 0.0)
        if sort_by == 'progress':
            return float(stats.get('pct_of_goal', 0.0) or 0.0)
        if sort_by == 'supporters':
            return int(stats.get('supporters', 0) or 0)
        if sort_by == 'recent30':
            return float(stats.get('raised_30d', 0.0) or 0.0)
        return getattr(page, 'created_at', datetime.min)

    pages.sort(key=_sort_key, reverse=(sort_dir == 'desc'))

    summary = {
        'total_pages': len(pages),
        'active_pages': sum(1 for p in pages if str(getattr(p, 'status', '') or '') == 'active'),
        'draft_pages': sum(1 for p in pages if str(getattr(p, 'status', '') or '') == 'draft'),
        'closed_pages': sum(1 for p in pages if str(getattr(p, 'status', '') or '') == 'closed'),
        'total_raised': round(total_raised, 2),
        'raised_30d': round(total_raised_30d, 2),
        'average_progress_pct': round((sum(float(page_stats[int(p.id)]['pct_of_goal']) for p in pages) / len(pages)), 1) if pages else 0.0,
        'pages_at_goal': pages_at_goal,
    }

    campaign_rollup_rows = sorted(
        campaign_rollups.values(),
        key=lambda row: float(row.get('raised', 0.0)),
        reverse=True,
    )
    owner_rollup_rows = sorted(
        owner_rollups.values(),
        key=lambda row: float(row.get('raised', 0.0)),
        reverse=True,
    )[:8]

    ai_context = {
        'active_page': 'p2p',
        'organization': org.name,
        'p2p_page_count': len(pages),
    }
    return render_template(
        'p2p_manage.html',
        form=form,
        pages=pages,
        page_stats=page_stats,
        summary=summary,
        donors=donors,
        campaign_rollups=campaign_rollup_rows,
        owner_rollups=owner_rollup_rows,
        automation_queue=automation_queue,
        filter_q=request.args.get('q', ''),
        filter_status=status_filter or '',
        filter_owner_id=owner_filter or 0,
        filter_campaign=request.args.get('campaign', ''),
        filter_sort_by=sort_by,
        filter_sort_dir=sort_dir,
        active_page='p2p',
        ai_context=ai_context,
    )


@main_bp.route('/donors')
@login_required
def donors_list():
    org = _current_org()
    query = request.args.get('q', '').strip()
    donor_type = request.args.get('donor_type', '').strip()

    donors = []
    if org:
        donors = DonorService().list_all_donors(org.id, donor_type=donor_type or None, search=query or None)
    delete_form = ConfirmDeleteForm()
    ai_context = {
        'active_page': 'donors',
        'organization': org.name if org else None,
        'donor_count': len(donors),
    }
    ctx = dict(
        donors=donors,
        delete_form=delete_form,
        active_page='donors',
        filter_q=query,
        filter_donor_type=donor_type,
        ai_context=ai_context,
    )
    if request.headers.get('HX-Request'):
        return render_template('_donors_rows.html', **ctx)
    return render_template('donors.html', **ctx)


@main_bp.route('/donors/<int:donor_id>', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'staff', 'viewer')
def donor_detail(donor_id: int):
    from ngo_homesuite.services.activity_timeline_service import ActivityTimelineService

    org = _current_org()
    donor_summary = ReportingService().donor_profile_summary(org.id, donor_id, recent_limit=10)
    donor = donor_summary['donor']
    donation_service = DonationService()

    quick_donation_form = DonorQuickDonationForm()
    quick_task_form = DonorQuickTaskForm(prefix='task')
    quick_donation_form.project_id.choices = [(0, 'General / None')] + [(p.id, p.name) for p in ProjectService().list_all_projects(org.id)]
    quick_donation_form.fund_id.choices = [(0, 'General / None')] + [
        (f.id, f.name)
        for f in FundService().list_funds(org.id, active_only=True, page=1, per_page=500)['items']
    ]

    if request.method == 'POST' and quick_task_form.submit.name in request.form:
        if quick_task_form.validate_on_submit():
            task = Task(
                organization_id=int(org.id),
                donor_id=int(donor.id),
                assigned_to_id=getattr(current_user, 'id', None),
                title=(quick_task_form.title.data or '').strip(),
                description=(quick_task_form.notes.data or '').strip() or None,
                task_type=str(quick_task_form.task_type.data or 'follow_up'),
                priority=str(quick_task_form.priority.data or 'medium'),
                status='open',
                due_date=quick_task_form.due_date.data,
            )
            db.session.add(task)
            db.session.commit()
            flash('Stewardship task created from donor profile.', 'success')
            return redirect(url_for('main.donor_detail', donor_id=donor.id))
        flash('Please fix task form errors and try again.', 'error')

    if request.method == 'POST' and quick_donation_form.submit.name in request.form and quick_donation_form.validate_on_submit():
        payment_method = str(quick_donation_form.payment_method.data or '').strip().lower()
        validation_errors: list[str] = []
        if payment_method == 'bank_transfer':
            if not str(quick_donation_form.bank_routing_number.data or '').strip():
                validation_errors.append('Bank routing number is required for bank transfers.')
            if not str(quick_donation_form.bank_account_number.data or '').strip():
                validation_errors.append('Bank account number is required for bank transfers.')
        elif payment_method == 'credit_card':
            if not str(quick_donation_form.card_holder_name.data or '').strip():
                validation_errors.append('Cardholder name is required for credit card donations.')
            if not str(quick_donation_form.card_last4.data or '').strip():
                validation_errors.append('Card last 4 digits are required for credit card donations.')

        if validation_errors:
            for message in validation_errors:
                flash(message, 'error')
        else:
            try:
                donation = donation_service.create_donation(
                    org_id=org.id,
                    donor_name=donor.name,
                    amount=float(quick_donation_form.amount.data),
                    currency=quick_donation_form.currency.data,
                    donor_email=donor.email,
                    donor_phone=donor.phone,
                    donor_id=donor.id,
                    project_id=quick_donation_form.project_id.data or None,
                    fund_id=quick_donation_form.fund_id.data or None,
                    payment_method=payment_method,
                    reference_number=_build_quick_donation_reference(quick_donation_form),
                    purpose=(quick_donation_form.purpose.data or '').strip() or None,
                    notes=_build_quick_donation_notes(quick_donation_form),
                    status='received',
                )
                donation_service.update_status(donation.id, org.id, 'processed', actor_id=getattr(current_user, 'id', None))
                if donor.email:
                    _issue_receipt_for_donation(donation, recipient_email=donor.email)
                flash('Donation recorded from donor profile.', 'success')
                return redirect(url_for('main.donor_detail', donor_id=donor.id))
            except ValueError as exc:
                flash(str(exc), 'error')
            except DonationConcurrencyError:
                flash('This donation was updated by another user. Please reload and try again.', 'error')

    ai_context = {
        'active_page': 'donors',
        'organization': org.name if org else None,
        'donor_id': donor.id,
        'donor_name': donor.name,
        'donation_count': donor_summary['donation_count'],
        'donation_total': donor_summary['donation_total'],
    }

    donor_ai_insights = None
    try:
        donor_ai_insights = CopilotToolRegistry().execute(
            "summarize_donor",
            {"donor_id": donor.id},
            {
                "organization_id": org.id,
                "actor": getattr(current_user, "username", "web"),
            },
        )
        if isinstance(donor_ai_insights, dict) and donor_ai_insights.get("error"):
            donor_ai_insights = None
    except Exception:
        donor_ai_insights = None

    activity_query = (request.args.get('activity_q') or '').strip()

    engagement_score = db.session.scalars(
        select(DonorEngagementScore).where(
            DonorEngagementScore.organization_id == int(org.id),
            DonorEngagementScore.donor_id == int(donor.id),
        ).limit(1)
    ).first()

    generated_tags: list[str] = []
    if donor.donor_type:
        generated_tags.append(f"type:{str(donor.donor_type).strip().lower()}")
    if getattr(donor, 'status', None):
        generated_tags.append(f"status:{str(donor.status).strip().lower()}")
    if getattr(donor, 'preferred_contact_method', None):
        generated_tags.append(f"contact:{str(donor.preferred_contact_method).strip().lower()}")
    if getattr(donor, 'source', None):
        generated_tags.append(f"source:{str(donor.source).strip().lower()}")
    if engagement_score and engagement_score.segment:
        generated_tags.append(f"segment:{str(engagement_score.segment).strip().lower()}")

    custom_field_schema = []
    org_metadata = org.metadata_json if isinstance(getattr(org, 'metadata_json', None), dict) else {}
    schema = org_metadata.get('custom_fields_schema', {}) if isinstance(org_metadata, dict) else {}
    if isinstance(schema, dict):
        donor_schema = schema.get('donor', [])
        if isinstance(donor_schema, list):
            custom_field_schema = [field for field in donor_schema if isinstance(field, dict)]

    donor_custom_metadata = getattr(donor, 'metadata_json', None)
    donor_custom_metadata = donor_custom_metadata if isinstance(donor_custom_metadata, dict) else {}
    donor_custom_field_values: list[dict[str, str]] = []
    for field in custom_field_schema:
        key = str(field.get('key') or '').strip()
        if not key:
            continue
        display_name = str(field.get('label') or key)
        value = donor_custom_metadata.get(key)
        if value is None and hasattr(donor, key):
            value = getattr(donor, key)
        if isinstance(value, (list, dict)):
            value_text = json.dumps(value, ensure_ascii=True)
        else:
            value_text = '' if value is None else str(value)
        donor_custom_field_values.append(
            {
                'key': key,
                'label': display_name,
                'value': value_text,
                'is_available': bool(value_text.strip()),
            }
        )

    household_name = None
    donor_relationships: list[dict[str, object]] = []
    try:
        conn = db.session.connection()
        household_row = conn.exec_driver_sql(
            """
            SELECT h.name
            FROM donors d
            LEFT JOIN households h ON h.id = d.household_id
            WHERE d.id = :donor_id AND d.organization_id = :org_id
            LIMIT 1
            """,
            {'donor_id': int(donor.id), 'org_id': int(org.id)},
        ).first()
        household_name = household_row[0] if household_row and household_row[0] else None

        relationship_rows = conn.exec_driver_sql(
            """
            SELECT
                dr.relationship_type,
                d2.id AS related_donor_id,
                d2.name AS related_donor_name
            FROM donor_relationships dr
            JOIN donors d1 ON d1.id = dr.from_donor_id
            JOIN donors d2 ON d2.id = dr.to_donor_id
            WHERE d1.id = :donor_id
              AND d1.organization_id = :org_id
            ORDER BY d2.name ASC
            """,
            {'donor_id': int(donor.id), 'org_id': int(org.id)},
        ).fetchall()
        donor_relationships = [
            {
                'relationship_type': row[0],
                'related_donor_id': row[1],
                'related_donor_name': row[2],
            }
            for row in relationship_rows
        ]
    except Exception as exc:
        current_app.logger.warning(
            'donor_detail_household_relationship_lookup_failed org_id=%s donor_id=%s error=%s',
            int(org.id),
            int(donor.id),
            exc,
            exc_info=True,
        )
        household_name = None
        donor_relationships = []

    stewardship_tasks = list(
        db.session.scalars(
            select(Task)
            .where(
                Task.organization_id == int(org.id),
                Task.donor_id == int(donor.id),
            )
            .order_by(Task.due_date.desc(), Task.created_at.desc())
            .limit(10)
        )
    )

    communication_history = list(
        db.session.scalars(
            select(CampaignEmailDelivery)
            .where(
                CampaignEmailDelivery.organization_id == int(org.id),
                CampaignEmailDelivery.donor_id == int(donor.id),
            )
            .order_by(CampaignEmailDelivery.sent_at.desc(), CampaignEmailDelivery.created_at.desc())
            .limit(10)
        )
    )

    # Fetch unified activity timeline
    timeline_items = []
    try:
        timeline_items = ActivityTimelineService.get_donor_timeline(
            org.id,
            donor_id,
            limit=25,
            search_query=activity_query or None,
        )
    except Exception:
        # If timeline fetch fails, continue without it
        timeline_items = []

    return render_template(
        'donor_detail.html',
        donor=donor,
        donation_count=donor_summary['donation_count'],
        donation_total=donor_summary['donation_total'],
        first_gift_date=donor_summary['first_gift_date'],
        last_gift_date=donor_summary['last_gift_date'],
        active_recurring_plans=donor_summary['active_recurring_plans'],
        recent_donations=donor_summary['recent_donations'],
        recurring_plans=donor_summary['recurring_plans'],
        timeline_items=timeline_items,
        activity_query=activity_query,
        active_page='donors',
        ai_context=ai_context,
        donor_ai_insights=donor_ai_insights,
        quick_donation_form=quick_donation_form,
        quick_task_form=quick_task_form,
        household_name=household_name,
        donor_relationships=donor_relationships,
        engagement_score=engagement_score,
        generated_tags=generated_tags,
        custom_field_schema=custom_field_schema,
        donor_custom_field_values=donor_custom_field_values,
        stewardship_tasks=stewardship_tasks,
        communication_history=communication_history,
    )


@main_bp.route('/donors/export/<string:file_type>')
@login_required
@require_step_up_auth
def donors_export(file_type: str):
    org = _current_org()
    if not org:
        flash('No organization available for export.', 'error')
        return redirect(url_for('main.donors_list'))

    query = request.args.get('q', '').strip()
    donor_type = request.args.get('donor_type', '').strip()
    donors = DonorService().list_all_donors(org.id, donor_type=donor_type or None, search=query or None)
    headers = ['ID', 'Name', 'Email', 'Phone', 'Type', 'Created At']
    rows = [
        [d.id, d.name, d.email or '', d.phone or '', d.donor_type, d.created_at.strftime('%Y-%m-%d %H:%M:%S') if d.created_at else '']
        for d in donors
    ]

    if file_type == 'csv':
        data = _build_csv_bytes(headers, rows)
        return send_file(BytesIO(data), mimetype='text/csv', as_attachment=True, download_name='donors.csv')
    if file_type == 'xlsx':
        data = _build_xlsx_bytes('Donors', headers, rows)
        return send_file(
            BytesIO(data),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='donors.xlsx',
        )
    flash('Unsupported export type.', 'error')
    return redirect(url_for('main.donors_list'))


@main_bp.route('/donors/new', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'staff')
def donor_create():
    org = _current_org()
    if not org:
        flash('No organization is available. Please seed data first.', 'error')
        return redirect(url_for('main.dashboard'))

    form = DonorForm()
    if request.method == 'GET':
        form.status.data = 'active'
        form.preferred_contact_method.data = 'email'
        form.communication_opt_in.data = True
    if form.validate_on_submit():
        donor_status = form.status.data or 'active'
        preferred_contact_method = form.preferred_contact_method.data or 'email'
        donor = DonorService().create_donor(
            org.id,
            form.name.data,
            salutation=form.salutation.data,
            preferred_name=form.preferred_name.data,
            email=form.email.data,
            phone=form.phone.data,
            donor_type=form.donor_type.data,
            status=donor_status,
            preferred_contact_method=preferred_contact_method,
            communication_opt_in=bool(form.communication_opt_in.data),
            address=form.address.data,
            city=form.city.data,
            country=form.country.data,
            postal_code=form.postal_code.data,
            employer=form.employer.data,
            source=form.source.data,
            notes=form.notes.data,
        )

        uploaded = request.files.get('photo')
        if uploaded is not None and uploaded.filename:
            try:
                donor.photo_path = _save_photo_upload(uploaded, entity='donors', org_id=org.id, record_id=int(donor.id))
                db.session.commit()
            except ValueError as exc:
                flash(str(exc), 'error')
                return redirect(url_for('main.donor_edit', donor_id=donor.id))

        flash('Donor created successfully.', 'success')
        return redirect(url_for('main.donors_list'))

    return render_template('donor_form.html', form=form, donor=None, is_edit=False, active_page='donors')


@main_bp.route('/donors/<int:donor_id>/edit', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'staff')
def donor_edit(donor_id: int):
    org = _current_org()
    donor = DonorService().get_donor(donor_id, org.id)
    form = DonorForm(obj=donor)

    if form.validate_on_submit():
        donor_status = form.status.data or 'active'
        preferred_contact_method = form.preferred_contact_method.data or 'email'
        donor = DonorService().update_donor(
            donor.id,
            org.id,
            name=form.name.data,
            salutation=form.salutation.data,
            preferred_name=form.preferred_name.data,
            email=form.email.data,
            phone=form.phone.data,
            donor_type=form.donor_type.data,
            status=donor_status,
            preferred_contact_method=preferred_contact_method,
            communication_opt_in=bool(form.communication_opt_in.data),
            address=form.address.data,
            city=form.city.data,
            country=form.country.data,
            postal_code=form.postal_code.data,
            employer=form.employer.data,
            source=form.source.data,
            notes=form.notes.data,
        )

        uploaded = request.files.get('photo')
        if uploaded is not None and uploaded.filename:
            donor.photo_path = _save_photo_upload(uploaded, entity='donors', org_id=org.id, record_id=int(donor.id))
            db.session.commit()

        flash('Donor updated successfully.', 'success')
        return redirect(url_for('main.donors_list'))

    return render_template('donor_form.html', form=form, is_edit=True, donor=donor, active_page='donors')


@main_bp.route('/media/donors/<int:donor_id>/photo')
@login_required
def donor_photo(donor_id: int):
    org = _current_org()
    if org is None:
        raise NotFound()

    donor = DonorService().get_donor(donor_id, org.id)
    if not donor.photo_path:
        raise NotFound()

    photo_path = _resolve_upload_path(str(donor.photo_path))
    if not photo_path.exists():
        raise NotFound()
    return send_file(photo_path)


@main_bp.route('/media/campaigns/<int:campaign_id>/photo')
@login_required
def campaign_photo(campaign_id: int):
    org = _current_org()
    if org is None:
        raise NotFound()

    campaign = db.session.scalars(
        select(Campaign).where(
            Campaign.id == int(campaign_id),
            Campaign.organization_id == int(org.id),
        ).limit(1)
    ).first()
    if campaign is None or not campaign.photo_path:
        raise NotFound()

    photo_path = _resolve_upload_path(str(campaign.photo_path))
    if not photo_path.exists():
        raise NotFound()
    return send_file(photo_path)


@main_bp.route('/donors/<int:donor_id>/delete', methods=['POST'])
@login_required
@roles_required('admin', 'staff')
@require_step_up_auth
def donor_delete(donor_id: int):
    org_id = getattr(current_user, 'organization_id', None)
    if org_id is None:
        flash('No organization is available.', 'error')
        return redirect(url_for('main.donors_list'))
    try:
        DonorService().delete_donor(donor_id, int(org_id))
    except ValueError:
        flash('Cannot delete donor with existing donations. Edit donor instead.', 'error')
        return redirect(url_for('main.donors_list'))
    flash('Donor deleted successfully.', 'success')
    return redirect(url_for('main.donors_list'))


@main_bp.route('/donors/dedupe')
@login_required
@roles_required('admin', 'staff')
def donor_dedupe():
    """Surface probable duplicate donors for manual merge."""
    org = _current_org()
    if not org:
        flash('No organization is available.', 'error')
        return redirect(url_for('main.donors_list'))

    donors = DonorService().list_all_donors(org.id)
    seen_email = {}
    seen_name_phone = {}
    candidates = []
    pair_keys = set()

    for donor in donors:
        email_key = _normalize_text(donor.email)
        name_phone_key = f"{_normalize_text(donor.name)}::{_normalize_phone(donor.phone)}"

        if email_key:
            first = seen_email.get(email_key)
            if first and first.id != donor.id:
                key = tuple(sorted([first.id, donor.id]))
                if key not in pair_keys:
                    pair_keys.add(key)
                    candidates.append({'primary': first, 'duplicate': donor, 'reason': 'Matching email'})
            else:
                seen_email[email_key] = donor

        # Name + phone is a fallback when email is missing.
        if _normalize_phone(donor.phone):
            first = seen_name_phone.get(name_phone_key)
            if first and first.id != donor.id:
                key = tuple(sorted([first.id, donor.id]))
                if key not in pair_keys:
                    pair_keys.add(key)
                    candidates.append({'primary': first, 'duplicate': donor, 'reason': 'Matching name + phone'})
            else:
                seen_name_phone[name_phone_key] = donor

    return render_template('donor_dedupe.html', candidates=candidates, active_page='donors')


@main_bp.route('/donors/import', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'staff')
def donor_import():
    """CSV/XLSX donor import with field mapping, preview, and duplicate suggestions."""
    org = _current_org()
    if not org:
        flash('No organization is available.', 'error')
        return redirect(url_for('main.donors_list'))

    headers: list[str] = []
    mapping = {'name': '', 'email': '', 'phone': '', 'donor_type': '', 'notes': ''}
    preview: dict[str, object] | None = None
    cache_id: str | None = None
    filename: str | None = None

    if request.method == 'POST':
        action = (request.form.get('action') or 'preview').strip().lower()

        if action == 'preview':
            uploaded = request.files.get('import_file')
            if uploaded is None or not uploaded.filename:
                flash('Please choose a CSV or XLSX file to preview.', 'error')
                return render_template('donor_import.html', active_page='donors')

            filename = uploaded.filename
            ext = Path(filename).suffix.lower()
            if ext not in _DONOR_IMPORT_ALLOWED_EXTENSIONS:
                flash('Unsupported file type. Use CSV or XLSX.', 'error')
                return render_template('donor_import.html', active_page='donors')

            cache_id = uuid.uuid4().hex
            cache_path = _donor_import_cache_dir() / f'{cache_id}{ext}'
            uploaded.save(cache_path)

            try:
                headers, rows = _parse_donor_import_file(cache_path)
            except Exception:
                cache_path.unlink(missing_ok=True)
                flash('Could not parse file. Please check the format and try again.', 'error')
                return render_template('donor_import.html', active_page='donors')

            if not headers:
                cache_path.unlink(missing_ok=True)
                flash('No header row found in file.', 'error')
                return render_template('donor_import.html', active_page='donors')

            if not rows:
                cache_path.unlink(missing_ok=True)
                flash('No data rows found in file.', 'error')
                return render_template('donor_import.html', active_page='donors')

            mapping = _extract_donor_import_mapping(request.form)
            guessed = _guess_donor_import_mapping(headers)
            for key, guessed_value in guessed.items():
                if not mapping.get(key):
                    mapping[key] = guessed_value

            preview = _build_donor_import_preview(org.id, rows, mapping)
            return render_template(
                'donor_import.html',
                active_page='donors',
                headers=headers,
                mapping=mapping,
                preview=preview,
                cache_id=cache_id,
                source_filename=filename,
            )

        if action == 'import':
            cache_id = (request.form.get('cache_id') or '').strip()
            if not cache_id:
                flash('Import preview expired. Please upload and preview again.', 'error')
                return redirect(url_for('main.donor_import'))

            cache_dir = _donor_import_cache_dir()
            candidates = list(cache_dir.glob(f'{cache_id}.*'))
            if not candidates:
                flash('Import preview cache not found. Please preview again.', 'error')
                return redirect(url_for('main.donor_import'))

            cache_path = candidates[0]
            filename = cache_path.name
            try:
                headers, rows = _parse_donor_import_file(cache_path)
            except Exception:
                cache_path.unlink(missing_ok=True)
                flash('Failed to read cached import file. Please preview again.', 'error')
                return redirect(url_for('main.donor_import'))

            mapping = _extract_donor_import_mapping(request.form)
            if not mapping.get('name'):
                flash('Please map a Name column before importing.', 'error')
                preview = _build_donor_import_preview(org.id, rows, mapping)
                return render_template(
                    'donor_import.html',
                    active_page='donors',
                    headers=headers,
                    mapping=mapping,
                    preview=preview,
                    cache_id=cache_id,
                    source_filename=filename,
                )

            preview = _build_donor_import_preview(org.id, rows, mapping)
            result = _apply_donor_import(org.id, preview['rows'])
            cache_path.unlink(missing_ok=True)

            flash(
                (
                    f"Import completed: created {result['created']}, "
                    f"skipped duplicates {result['skipped_duplicates']}, "
                    f"skipped errors {result['skipped_errors']}."
                ),
                'success',
            )
            return redirect(url_for('main.donors_list'))

    return render_template('donor_import.html', active_page='donors')


@main_bp.route('/donors/merge', methods=['POST'])
@login_required
@roles_required('admin', 'staff')
def donor_merge():
    """Merge a duplicate donor into a primary donor."""
    org = _current_org()
    if not org:
        flash('No organization is available.', 'error')
        return redirect(url_for('main.donors_list'))

    primary_id = request.form.get('primary_id', type=int)
    duplicate_id = request.form.get('duplicate_id', type=int)
    if not primary_id or not duplicate_id or primary_id == duplicate_id:
        flash('Invalid merge request.', 'error')
        return redirect(url_for('main.donor_dedupe'))

    try:
        DonorService().merge_donors(org.id, primary_id, duplicate_id)
    except ValueError:
        flash('Invalid merge request.', 'error')
        return redirect(url_for('main.donor_dedupe'))
    flash('Duplicate donor merged successfully.', 'success')
    return redirect(url_for('main.donors_list'))


@main_bp.route('/donations/new', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'staff')
def donation_create():
    org = _current_org()
    if not org:
        flash('No organization is available. Please seed data first.', 'error')
        return redirect(url_for('main.dashboard'))

    donor_service = DonorService()
    donation_service = DonationService()
    donor_options = [(0, 'Select a donor')] + [(d.id, d.name) for d in donor_service.list_all_donors(org.id)]
    campaign_options = [(0, 'General / None')] + [
        (c.id, c.name)
        for c in db.session.scalars(
            select(Campaign).where(Campaign.organization_id == org.id).order_by(Campaign.name.asc())
        ).all()
    ]
    project_options = [(0, 'General / None')] + [(p.id, p.name) for p in ProjectService().list_all_projects(org.id)]
    fund_options = [(0, 'General / None')] + [
        (f.id, f.name)
        for f in FundService().list_funds(org.id, active_only=True, page=1, per_page=500)['items']
    ]

    form = DonationForm()
    form.donor_id.choices = donor_options
    form.campaign_id.choices = campaign_options
    form.project_id.choices = project_options
    form.fund_id.choices = fund_options

    donor_id_query = request.args.get('donor_id', type=int)
    if donor_id_query:
        form.donor_id.data = donor_id_query

    if form.validate_on_submit():
        try:
            donor = donor_service.get_donor(form.donor_id.data, org.id)
        except DonorNotFound:
            donor = None
        if donor is None:
            flash('Please select a valid donor.', 'error')
            return render_template('donation_form.html', form=form, active_page='donations')

        try:
            donation = donation_service.create_donation(
                org_id=org.id,
                donor_name=donor.name,
                amount=form.amount.data,
                currency=form.currency.data,
                donor_email=donor.email,
                donor_phone=donor.phone,
                donor_id=donor.id,
                campaign_id=form.campaign_id.data or None,
                project_id=form.project_id.data or None,
                fund_id=form.fund_id.data or None,
                payment_method=form.payment_method.data,
                channel=form.channel.data or None,
                reference_number=_build_quick_donation_reference(form),
                purpose=form.purpose.data,
                is_anonymous=bool(form.is_anonymous.data),
                public_display_name=form.public_display_name.data,
                tribute_type=form.tribute_type.data or None,
                tribute_honoree_name=form.tribute_honoree_name.data,
                tribute_honoree_contact=form.tribute_honoree_contact.data,
                soft_credit_name=form.soft_credit_name.data,
                notes=_build_quick_donation_notes(form),
                status='received',
            )
            donation_service.update_status(donation.id, org.id, 'processed', actor_id=getattr(current_user, 'id', None))
            _issue_receipt_for_donation(donation, recipient_email=donor.email)
        except ValueError as exc:
            flash(str(exc), 'error')
            return render_template('donation_form.html', form=form, active_page='donations')
        except DonationConcurrencyError:
            flash('This donation was updated by another user. Please reload and try again.', 'error')
            return render_template('donation_form.html', form=form, active_page='donations')
        flash('Donation recorded successfully and receipt generated.', 'success')
        return redirect(url_for('main.dashboard'))

    return render_template('donation_form.html', form=form, active_page='donations')


@main_bp.route('/donations/recurring', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'staff')
def recurring_donations():
    """Create and view recurring donation plans."""
    org = _current_org()
    if not org:
        flash('No organization is available.', 'error')
        return redirect(url_for('main.dashboard'))

    donor_service = DonorService()
    donation_service = DonationService()
    donor_options = [(0, 'Select a donor')] + [
        (d.id, f"{d.name} ({d.email or 'no email'})")
        for d in donor_service.list_all_donors(org.id)
    ]

    form = RecurringDonationForm()
    form.donor_id.choices = donor_options

    if form.validate_on_submit():
        try:
            donor = donor_service.get_donor(form.donor_id.data, org.id)
        except DonorNotFound:
            donor = None
        if donor is None:
            flash('Please select a valid donor.', 'error')
            return render_template('recurring_donations.html', form=form, plans=[], active_page='donations')

        try:
            donation_service.create_recurring_plan(
                org.id,
                donor.id,
                amount=form.amount.data,
                currency=form.currency.data,
                payment_method=form.payment_method.data,
                purpose=form.purpose.data,
                frequency=form.frequency.data,
                next_charge_date=_next_charge_date(date.today(), form.frequency.data),
            )
        except ValueError as exc:
            flash(str(exc), 'error')
            plans = donation_service.list_recurring_plans(org.id)
            return render_template('recurring_donations.html', form=form, plans=plans, active_page='donations')
        except DonationConcurrencyError:
            flash('Recurring plan changed while saving. Please try again.', 'error')
            plans = donation_service.list_recurring_plans(org.id)
            return render_template('recurring_donations.html', form=form, plans=plans, active_page='donations')
        flash('Recurring donation plan created.', 'success')
        return redirect(url_for('main.recurring_donations'))

    plans = donation_service.list_recurring_plans(org.id)
    return render_template('recurring_donations.html', form=form, plans=plans, active_page='donations')


@main_bp.route('/donations/recurring/process', methods=['POST'])
@login_required
@roles_required('admin', 'staff')
def process_recurring_donations():
    """Run due recurring plans and handle failures for missing payment contact data."""
    org = _current_org()
    if not org:
        flash('No organization is available.', 'error')
        return redirect(url_for('main.dashboard'))

    result = DonationService().process_due_recurring_plans(org.id, run_date=date.today())
    processed = result.get('processed', 0)
    failed = result.get('failed', 0)
    flash(f'Recurring processing complete: {processed} processed, {failed} failed.', 'info')
    return redirect(url_for('main.recurring_donations'))


@main_bp.route('/donations')
@login_required
def donations_list():
    org = _current_org()
    query = request.args.get('q', '').strip()
    fund_id = request.args.get('fund_id', type=int)
    payment_method = request.args.get('payment_method', '').strip()
    channel = request.args.get('channel', '').strip().lower()
    status = request.args.get('status', '').strip()
    currency = request.args.get('currency', '').strip().upper()
    donor_type = request.args.get('donor_type', '').strip().lower()
    start_date = _parse_date(request.args.get('start_date', ''))
    end_date = _parse_date(request.args.get('end_date', ''))
    sort_by = request.args.get('sort_by', 'date').strip().lower()
    sort_dir = request.args.get('sort_dir', 'desc').strip().lower()
    if sort_by not in {'date', 'amount', 'donor', 'status'}:
        sort_by = 'date'
    if sort_dir not in {'asc', 'desc'}:
        sort_dir = 'desc'

    per_page_options = [25, 50, 100, 250]
    per_page = request.args.get('per_page', type=int) or 50
    if per_page not in per_page_options:
        per_page = 50
    page = request.args.get('page', type=int) or 1
    page = max(page, 1)

    min_amount = _parse_float(request.args.get('min_amount', ''))
    max_amount = _parse_float(request.args.get('max_amount', ''))

    donations = []
    if org:
        donations = DonationService().list_filtered_donations(
            org.id,
            search=query or None,
            payment_method=payment_method or None,
            status=status or None,
            min_amount=min_amount,
            max_amount=max_amount,
        )

    if currency:
        donations = [d for d in donations if str(getattr(d, 'currency', '') or '').upper() == currency]

    if fund_id:
        donations = [d for d in donations if int(getattr(d, 'fund_id', 0) or 0) == int(fund_id)]

    if channel:
        donations = [
            d for d in donations
            if str(getattr(d, 'channel', '') or '').strip().lower() == channel
        ]

    if donor_type:
        donations = [
            d for d in donations
            if str(getattr(getattr(d, 'donor', None), 'donor_type', '') or '').strip().lower() == donor_type
        ]

    if start_date or end_date:
        filtered: list[Donation] = []
        for d in donations:
            donation_dt = getattr(d, 'donation_date', None)
            donation_day = donation_dt.date() if hasattr(donation_dt, 'date') else None
            if donation_day is None:
                continue
            if start_date and donation_day < start_date:
                continue
            if end_date and donation_day > end_date:
                continue
            filtered.append(d)
        donations = filtered

    def _safe_amount(item: Donation) -> float:
        try:
            return float(getattr(item, 'amount', 0) or 0)
        except (TypeError, ValueError):
            return 0.0

    def _sort_key(item: Donation):
        if sort_by == 'amount':
            return _safe_amount(item)
        if sort_by == 'donor':
            return str(getattr(item, 'donor_name', '') or '').strip().lower()
        if sort_by == 'status':
            return str(getattr(item, 'status', '') or '').strip().lower()
        return getattr(item, 'donation_date', datetime.min)

    donations.sort(key=_sort_key, reverse=(sort_dir == 'desc'))

    total_filtered = len(donations)
    total_pages = max(1, (total_filtered + per_page - 1) // per_page)
    if page > total_pages:
        page = total_pages
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paged_donations = donations[start_idx:end_idx]

    total_visible_amount = 0.0
    for donation in donations:
        try:
            total_visible_amount += float(getattr(donation, 'amount', 0) or 0)
        except (TypeError, ValueError):
            continue

    average_visible_amount = (total_visible_amount / total_filtered) if total_filtered else 0.0
    status_counts = {
        'received': 0,
        'processed': 0,
        'receipted': 0,
        'failed': 0,
        'refunded': 0,
    }
    for d in donations:
        status_key = str(getattr(d, 'status', '') or '').strip().lower()
        if status_key in status_counts:
            status_counts[status_key] += 1

    row_action_form = DonationRowActionForm()
    bulk_action_form = BulkDonationActionForm()
    next_status_map = {
        'pending': ['received', 'failed'],
        'received': ['processed', 'refunded'],
        'processed': ['receipted', 'refunded'],
        'receipted': ['refunded'],
        'failed': ['pending'],
        'refunded': [],
    }
    donation_next_statuses: dict[int, list[str]] = {
        int(d.id): next_status_map.get(str(getattr(d, 'status', '') or '').strip().lower(), [])
        for d in paged_donations
    }

    available_currencies = sorted(
        {
            str(getattr(d, 'currency', '') or '').upper()
            for d in donations
            if str(getattr(d, 'currency', '') or '').strip()
        }
    )
    available_funds = []
    if org:
        available_funds = FundService().list_all_funds(org.id)
    available_channels = sorted(
        {
            str(getattr(d, 'channel', '') or '').strip().lower()
            for d in donations
            if str(getattr(d, 'channel', '') or '').strip()
        }
    )
    donor_type_options = ['individual', 'corporate', 'foundation', 'anonymous']

    ai_context = {
        'active_page': 'donations',
        'organization': org.name if org else None,
        'donation_count': total_filtered,
        'total_donations': total_visible_amount,
        'average_donation': average_visible_amount,
    }
    return render_template(
        'donations.html',
        donations=paged_donations,
        total_filtered=total_filtered,
        status_counts=status_counts,
        row_action_form=row_action_form,
        bulk_action_form=bulk_action_form,
        donation_next_statuses=donation_next_statuses,
        page=page,
        total_pages=total_pages,
        per_page=per_page,
        per_page_options=per_page_options,
        available_currencies=available_currencies,
        available_funds=available_funds,
        available_channels=available_channels,
        donor_type_options=donor_type_options,
        active_page='donations',
        filter_q=query,
        filter_fund_id=fund_id,
        filter_payment_method=payment_method,
        filter_channel=channel,
        filter_status=status,
        filter_currency=currency,
        filter_donor_type=donor_type,
        filter_start_date=request.args.get('start_date', ''),
        filter_end_date=request.args.get('end_date', ''),
        filter_sort_by=sort_by,
        filter_sort_dir=sort_dir,
        filter_min_amount=request.args.get('min_amount', ''),
        filter_max_amount=request.args.get('max_amount', ''),
        ai_context=ai_context,
    )


@main_bp.route('/donations/<int:donation_id>/status', methods=['POST'])
@login_required
@roles_required('admin', 'staff')
def donation_status_update(donation_id: int):
    org = _current_org()
    if not org:
        flash('No organization is available.', 'error')
        return redirect(url_for('main.donations_list'))

    form = DonationRowActionForm()
    if not form.validate_on_submit():
        flash('Invalid donation status update request.', 'error')
        return redirect(url_for('main.donations_list'))

    if int(form.donation_id.data or 0) != int(donation_id):
        flash('Donation status update request did not match record id.', 'error')
        return redirect(url_for('main.donations_list'))

    new_status = str(form.new_status.data or '').strip().lower()
    if not new_status:
        flash('Please choose a valid status.', 'error')
        return redirect(url_for('main.donations_list'))

    svc = DonationService()
    try:
        donation = svc.update_status(donation_id, org.id, new_status, actor_id=getattr(current_user, 'id', None))
        if new_status == 'receipted':
            _issue_receipt_for_donation(donation, recipient_email=donation.donor_email)
        flash(f'Donation #{donation_id} updated to {new_status}.', 'success')
    except DonationNotFound:
        flash('Donation not found.', 'error')
    except InvalidStatusTransition as exc:
        flash(str(exc), 'error')
    except DonationConcurrencyError:
        flash('This donation changed while updating. Please retry.', 'error')
    except ValueError as exc:
        flash(str(exc), 'error')

    next_url = str(form.next_url.data or '').strip()
    if next_url.startswith('/'):
        return redirect(next_url)
    return redirect(url_for('main.donations_list'))


@main_bp.route('/donations/<int:donation_id>/receipt/resend', methods=['POST'])
@login_required
@roles_required('admin', 'staff')
def donation_receipt_resend(donation_id: int):
    org = _current_org()
    if not org:
        flash('No organization is available.', 'error')
        return redirect(url_for('main.donations_list'))

    form = DonationRowActionForm()
    if not form.validate_on_submit():
        flash('Invalid receipt resend request.', 'error')
        return redirect(url_for('main.donations_list'))

    if int(form.donation_id.data or 0) != int(donation_id):
        flash('Receipt resend request did not match record id.', 'error')
        return redirect(url_for('main.donations_list'))

    try:
        donation = DonationService().get_donation(donation_id, org.id)
        recipient_email = str(getattr(donation, 'donor_email', '') or '').strip() or None
        _issue_receipt_for_donation(donation, recipient_email=recipient_email)
        flash(f'Receipt regenerated for donation #{donation_id}.', 'success')
    except DonationNotFound:
        flash('Donation not found.', 'error')
    except ValueError as exc:
        flash(str(exc), 'error')

    next_url = str(form.next_url.data or '').strip()
    if next_url.startswith('/'):
        return redirect(next_url)
    return redirect(url_for('main.donations_list'))


@main_bp.route('/donations/bulk/status', methods=['POST'])
@login_required
@roles_required('admin', 'staff')
def donations_bulk_status_update():
    org = _current_org()
    if not org:
        flash('No organization is available.', 'error')
        return redirect(url_for('main.donations_list'))

    form = BulkDonationActionForm()
    if not form.validate_on_submit():
        flash('Invalid bulk status update request.', 'error')
        return redirect(url_for('main.donations_list'))

    selected_ids = [
        int(raw_id)
        for raw_id in request.form.getlist('donation_ids')
        if str(raw_id or '').strip().isdigit()
    ]
    if not selected_ids:
        flash('Select at least one donation for bulk status update.', 'error')
        return redirect(url_for('main.donations_list'))

    new_status = str(form.new_status.data or '').strip().lower()
    if not new_status:
        flash('Choose a status to apply.', 'error')
        return redirect(url_for('main.donations_list'))

    svc = DonationService()
    updated = 0
    skipped = 0
    for donation_id in selected_ids:
        try:
            donation = svc.update_status(donation_id, org.id, new_status, actor_id=getattr(current_user, 'id', None))
            if new_status == 'receipted':
                _issue_receipt_for_donation(donation, recipient_email=donation.donor_email)
            updated += 1
        except (DonationNotFound, InvalidStatusTransition, DonationConcurrencyError, ValueError):
            skipped += 1

    flash(f'Bulk status update complete: {updated} updated, {skipped} skipped.', 'success' if updated else 'info')
    next_url = str(form.next_url.data or '').strip()
    if next_url.startswith('/'):
        return redirect(next_url)
    return redirect(url_for('main.donations_list'))


@main_bp.route('/donations/bulk/receipt/resend', methods=['POST'])
@login_required
@roles_required('admin', 'staff')
def donations_bulk_receipt_resend():
    org = _current_org()
    if not org:
        flash('No organization is available.', 'error')
        return redirect(url_for('main.donations_list'))

    form = BulkDonationActionForm()
    if not form.validate_on_submit():
        flash('Invalid bulk receipt resend request.', 'error')
        return redirect(url_for('main.donations_list'))

    selected_ids = [
        int(raw_id)
        for raw_id in request.form.getlist('donation_ids')
        if str(raw_id or '').strip().isdigit()
    ]
    if not selected_ids:
        flash('Select at least one donation to resend receipts.', 'error')
        return redirect(url_for('main.donations_list'))

    resent = 0
    skipped = 0
    for donation_id in selected_ids:
        try:
            donation = DonationService().get_donation(donation_id, org.id)
            recipient_email = str(getattr(donation, 'donor_email', '') or '').strip() or None
            _issue_receipt_for_donation(donation, recipient_email=recipient_email)
            resent += 1
        except (DonationNotFound, ValueError):
            skipped += 1

    flash(f'Bulk receipt resend complete: {resent} resent, {skipped} skipped.', 'success' if resent else 'info')
    next_url = str(form.next_url.data or '').strip()
    if next_url.startswith('/'):
        return redirect(next_url)
    return redirect(url_for('main.donations_list'))


@main_bp.route('/donations/export/<string:file_type>')
@login_required
@require_step_up_auth
def donations_export(file_type: str):
    org = _current_org()
    if not org:
        flash('No organization available for export.', 'error')
        return redirect(url_for('main.donations_list'))

    query = request.args.get('q', '').strip()
    fund_id = request.args.get('fund_id', type=int)
    payment_method = request.args.get('payment_method', '').strip()
    channel = request.args.get('channel', '').strip().lower()
    status = request.args.get('status', '').strip()
    min_amount = _parse_float(request.args.get('min_amount', ''))
    max_amount = _parse_float(request.args.get('max_amount', ''))

    donations = DonationService().list_filtered_donations(
        org.id,
        search=query or None,
        payment_method=payment_method or None,
        status=status or None,
        min_amount=min_amount,
        max_amount=max_amount,
    )
    if channel:
        donations = [
            d for d in donations
            if str(getattr(d, 'channel', '') or '').strip().lower() == channel
        ]
    if fund_id:
        donations = [d for d in donations if int(getattr(d, 'fund_id', 0) or 0) == int(fund_id)]
    headers = ['ID', 'Date', 'Donor', 'Email', 'Amount', 'Currency', 'Status', 'Payment Method', 'Purpose', 'Reference']
    rows = [
        [
            d.id,
            d.donation_date.strftime('%Y-%m-%d') if d.donation_date else '',
            d.donor_name,
            d.donor_email or '',
            d.amount,
            d.currency,
            d.status,
            d.payment_method or '',
            d.purpose or '',
            d.reference_number or '',
        ]
        for d in donations
    ]

    if file_type == 'csv':
        data = _build_csv_bytes(headers, rows)
        return send_file(BytesIO(data), mimetype='text/csv', as_attachment=True, download_name='donations.csv')
    if file_type == 'xlsx':
        data = _build_xlsx_bytes('Donations', headers, rows)
        return send_file(
            BytesIO(data),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='donations.xlsx',
        )
    if file_type == 'iif':
        iif_rows = [
            [
                d.id,
                d.donation_date.strftime('%Y-%m-%d') if d.donation_date else '',
                d.donor_name or 'Unknown Donor',
                d.amount,
                d.purpose or 'Donation',
            ]
            for d in donations
        ]
        data = _build_iif_bytes('DEPOSIT', iif_rows)
        return send_file(BytesIO(data), mimetype='text/plain', as_attachment=True, download_name='donations.iif')
    flash('Unsupported export type.', 'error')
    return redirect(url_for('main.donations_list'))


@main_bp.route('/donations/<int:donation_id>/receipt')
@login_required
def donation_receipt(donation_id: int):
    org = _current_org()
    try:
        donation = DonationService().get_donation(donation_id, org.id)
    except DonationNotFound:
        abort(404)
    donor = donation.donor

    donation_payload = {
        'amount_cents': int(round(float(donation.amount) * 100)),
        'currency': donation.currency,
        'received_at': donation.donation_date.strftime('%Y-%m-%d'),
    }
    donor_payload = {
        'name': donor.name if donor else donation.donor_name,
        'address': '',
    }

    pdf_data = generate_receipt_pdf_bytes(donation_payload, donor_payload)
    file_name = f"receipt-{donation.id}.pdf"
    return send_file(
        BytesIO(pdf_data),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=file_name,
    )


@main_bp.route('/expenses')
@login_required
def expenses_list():
    org = _current_org()
    query = request.args.get('q', '').strip()
    fund_id = request.args.get('fund_id', type=int)
    min_amount = _parse_float(request.args.get('min_amount', ''))
    max_amount = _parse_float(request.args.get('max_amount', ''))
    available_funds = []

    expenses = []
    if org:
        available_funds = FundService().list_all_funds(org.id)
        expenses = ExpenseService().list_filtered_expenses(
            org.id,
            search=query or None,
            min_amount=min_amount,
            max_amount=max_amount,
        )

    if fund_id:
        expenses = [e for e in expenses if int(getattr(e, 'fund_id', 0) or 0) == int(fund_id)]
    ai_context = {
        'active_page': 'expenses',
        'organization': org.name if org else None,
        'expense_count': len(expenses),
        'total_expenses': sum(float(e.amount or 0) for e in expenses),
    }
    return render_template(
        'expenses.html',
        expenses=expenses,
        available_funds=available_funds,
        active_page='expenses',
        filter_q=query,
        filter_fund_id=fund_id,
        filter_min_amount=request.args.get('min_amount', ''),
        filter_max_amount=request.args.get('max_amount', ''),
        ai_context=ai_context,
    )


@main_bp.route('/expenses/new', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'staff')
def expense_create():
    org = _current_org()
    if not org:
        flash('No organization is available. Please seed data first.', 'error')
        return redirect(url_for('main.dashboard'))

    project_options = [(0, 'General / None')] + [(p.id, p.name) for p in ProjectService().list_all_projects(org.id)]
    fund_options = [(0, 'General / None')] + [
        (f.id, f.name)
        for f in FundService().list_all_funds(org.id, active_only=True)
    ]

    form = ExpenseForm()
    form.project_id.choices = project_options
    form.fund_id.choices = fund_options

    if form.validate_on_submit():
        try:
            ExpenseService().create_expense(
                org.id,
                project_id=form.project_id.data or None,
                fund_id=form.fund_id.data or None,
                amount=form.amount.data,
                currency=form.currency.data,
                payee=form.payee.data,
                description=form.description.data,
            )
        except ValueError as exc:
            flash(f'Could not record expense: {exc}', 'error')
            return render_template('expense_form.html', form=form, active_page='expenses')
        flash('Expense recorded successfully.', 'success')
        return redirect(url_for('main.expenses_list'))

    return render_template('expense_form.html', form=form, active_page='expenses')


@main_bp.route('/expenses/export/<string:file_type>')
@login_required
@require_step_up_auth
def expenses_export(file_type: str):
    org = _current_org()
    if not org:
        flash('No organization available for export.', 'error')
        return redirect(url_for('main.expenses_list'))

    query = request.args.get('q', '').strip()
    fund_id = request.args.get('fund_id', type=int)
    min_amount = _parse_float(request.args.get('min_amount', ''))
    max_amount = _parse_float(request.args.get('max_amount', ''))

    expenses = ExpenseService().list_filtered_expenses(
        org.id,
        search=query or None,
        min_amount=min_amount,
        max_amount=max_amount,
    )
    if fund_id:
        expenses = [e for e in expenses if int(getattr(e, 'fund_id', 0) or 0) == int(fund_id)]
    headers = ['ID', 'Date', 'Payee', 'Amount', 'Currency', 'Project', 'Fund', 'Description']
    rows = [
        [
            e.id,
            e.paid_at.strftime('%Y-%m-%d') if e.paid_at else '',
            e.payee or '',
            e.amount,
            e.currency,
            e.project.name if e.project else '',
            e.fund.name if e.fund else '',
            e.description or '',
        ]
        for e in expenses
    ]

    if file_type == 'csv':
        data = _build_csv_bytes(headers, rows)
        return send_file(BytesIO(data), mimetype='text/csv', as_attachment=True, download_name='expenses.csv')
    if file_type == 'xlsx':
        data = _build_xlsx_bytes('Expenses', headers, rows)
        return send_file(
            BytesIO(data),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='expenses.xlsx',
        )
    if file_type == 'iif':
        iif_rows = [
            [
                e.id,
                e.paid_at.strftime('%Y-%m-%d') if e.paid_at else '',
                (e.payee or e.description or 'Expense'),
                e.amount,
                e.description or 'Expense',
            ]
            for e in expenses
        ]
        data = _build_iif_bytes('CHECK', iif_rows)
        return send_file(BytesIO(data), mimetype='text/plain', as_attachment=True, download_name='expenses.iif')
    flash('Unsupported export type.', 'error')
    return redirect(url_for('main.expenses_list'))


@main_bp.route('/projects')
@login_required
def projects_dashboard():
    org = _current_org()
    query = request.args.get('q', '').strip()
    status = request.args.get('status', '').strip()

    projects = []
    if org:
        projects = ProjectService().list_all_projects(org.id, search=query or None, status=status or None)
    ai_context = {
        'active_page': 'projects',
        'organization': org.name if org else None,
        'project_count': len(projects),
    }
    return render_template(
        'projects.html',
        projects=projects,
        active_page='projects',
        filter_q=query,
        filter_status=status,
        ai_context=ai_context,
    )


@main_bp.route('/projects/export/<string:file_type>')
@login_required
@require_step_up_auth
def projects_export(file_type: str):
    org = _current_org()
    if not org:
        flash('No organization available for export.', 'error')
        return redirect(url_for('main.projects_dashboard'))

    query = request.args.get('q', '').strip()
    status = request.args.get('status', '').strip()
    projects = ProjectService().list_all_projects(org.id, search=query or None, status=status or None)
    headers = ['ID', 'Name', 'Program', 'Status', 'Currency', 'Budget', 'Spent']
    rows = [[p.id, p.name, p.program or '', p.status, p.currency, p.budget, p.spent] for p in projects]

    if file_type == 'csv':
        data = _build_csv_bytes(headers, rows)
        return send_file(BytesIO(data), mimetype='text/csv', as_attachment=True, download_name='projects.csv')
    if file_type == 'xlsx':
        data = _build_xlsx_bytes('Projects', headers, rows)
        return send_file(
            BytesIO(data),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='projects.xlsx',
        )
    flash('Unsupported export type.', 'error')
    return redirect(url_for('main.projects_dashboard'))


@main_bp.route('/projects/new', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'staff')
def project_create():
    org = _current_org()
    if not org:
        flash('No organization is available. Please seed data first.', 'error')
        return redirect(url_for('main.dashboard'))

    form = ProjectForm()
    if form.validate_on_submit():
        ProjectService().create_project(
            org.id,
            name=form.name.data,
            description=form.description.data,
            program=form.program.data,
            budget=form.budget.data,
            spent=form.spent.data,
            currency=form.currency.data,
            status=form.status.data,
        )
        flash('Project created successfully.', 'success')
        return redirect(url_for('main.projects_dashboard'))

    return render_template('project_form.html', form=form, is_edit=False, active_page='projects')


@main_bp.route('/projects/<int:project_id>/edit', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'staff')
def project_edit(project_id: int):
    org = _current_org()
    try:
        project = ProjectService().get_project(project_id, org.id)
    except ProjectNotFound:
        abort(404)
    form = ProjectForm(obj=project)
    if form.validate_on_submit():
        ProjectService().update_project(
            project.id,
            org.id,
            name=form.name.data,
            description=form.description.data,
            program=form.program.data,
            budget=form.budget.data,
            spent=form.spent.data,
            currency=form.currency.data,
            status=form.status.data,
        )
        flash('Project updated successfully.', 'success')
        return redirect(url_for('main.projects_dashboard'))

    return render_template('project_form.html', form=form, is_edit=True, project=project, active_page='projects')


@main_bp.route('/funds')
@login_required
def funds_list():
    org = _current_org()
    query = request.args.get('q', '').strip()
    status = request.args.get('status', '').strip()
    has_activity = request.args.get('has_activity', '').strip().lower()
    trend_window = request.args.get('trend_window', '6m').strip().lower()
    sort_by = request.args.get('sort_by', 'name').strip().lower()
    sort_dir = request.args.get('sort_dir', 'asc').strip().lower()
    if sort_by not in {'name', 'status', 'updated', 'in', 'out', 'net', 'donations', 'expenses'}:
        sort_by = 'name'
    if sort_dir not in {'asc', 'desc'}:
        sort_dir = 'asc'
    if trend_window not in {'30d', '90d', 'ytd', '6m'}:
        trend_window = '6m'

    per_page_options = [25, 50, 100, 250]
    per_page = request.args.get('per_page', type=int) or 50
    if per_page not in per_page_options:
        per_page = 50
    page = request.args.get('page', type=int) or 1
    page = max(page, 1)

    funds = []
    fund_metrics: dict[int, dict[str, float | int]] = {}
    summary = {
        'total_funds': 0,
        'active_funds': 0,
        'inactive_funds': 0,
        'total_in': 0.0,
        'total_out': 0.0,
        'net': 0.0,
    }
    if org:
        funds = FundService().list_all_funds(org.id, search=query or None, status=status or None)
        fund_ids = [int(f.id) for f in funds]
        summary['total_funds'] = len(funds)
        summary['active_funds'] = sum(1 for f in funds if bool(getattr(f, 'is_active', False)))
        summary['inactive_funds'] = max(0, int(summary['total_funds']) - int(summary['active_funds']))

        donations_by_fund: dict[int, dict[str, float | int]] = {}
        expenses_by_fund: dict[int, dict[str, float | int]] = {}

        if fund_ids:
            donation_rollups = (
                db.session.query(
                    Donation.fund_id,
                    func.count(Donation.id),
                    func.coalesce(func.sum(Donation.amount), 0.0),
                )
                .filter(
                    Donation.organization_id == org.id,
                    Donation.fund_id.in_(fund_ids),
                )
                .group_by(Donation.fund_id)
                .all()
            )
            for fund_id, count, amount in donation_rollups:
                if fund_id is None:
                    continue
                donations_by_fund[int(fund_id)] = {
                    'count': int(count or 0),
                    'amount': float(amount or 0.0),
                }

            expense_rollups = (
                db.session.query(
                    Expense.fund_id,
                    func.count(Expense.id),
                    func.coalesce(func.sum(Expense.amount), 0.0),
                )
                .filter(
                    Expense.organization_id == org.id,
                    Expense.fund_id.in_(fund_ids),
                )
                .group_by(Expense.fund_id)
                .all()
            )
            for fund_id, count, amount in expense_rollups:
                if fund_id is None:
                    continue
                expenses_by_fund[int(fund_id)] = {
                    'count': int(count or 0),
                    'amount': float(amount or 0.0),
                }

            now_dt = datetime.now(timezone.utc).replace(tzinfo=None)
            trend_labels: list[str] = []
            trend_ranges: list[tuple[datetime, datetime]] = []

            if trend_window == '30d':
                start_dt = (now_dt - timedelta(days=29)).replace(hour=0, minute=0, second=0, microsecond=0)
                for idx in range(6):
                    bucket_start = start_dt + timedelta(days=idx * 5)
                    bucket_end = bucket_start + timedelta(days=4, hours=23, minutes=59, seconds=59)
                    trend_labels.append(bucket_start.strftime('%m/%d'))
                    trend_ranges.append((bucket_start, bucket_end))
            elif trend_window == '90d':
                start_dt = (now_dt - timedelta(days=89)).replace(hour=0, minute=0, second=0, microsecond=0)
                for idx in range(6):
                    bucket_start = start_dt + timedelta(days=idx * 15)
                    bucket_end = bucket_start + timedelta(days=14, hours=23, minutes=59, seconds=59)
                    trend_labels.append(bucket_start.strftime('%m/%d'))
                    trend_ranges.append((bucket_start, bucket_end))
            elif trend_window == 'ytd':
                for month in range(1, now_dt.month + 1):
                    bucket_start = datetime(now_dt.year, month, 1)
                    if month == 12:
                        next_month = datetime(now_dt.year + 1, 1, 1)
                    else:
                        next_month = datetime(now_dt.year, month + 1, 1)
                    bucket_end = next_month - timedelta(seconds=1)
                    trend_labels.append(bucket_start.strftime('%b'))
                    trend_ranges.append((bucket_start, bucket_end))
            else:
                month_anchor = now_dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                offsets: list[tuple[int, int]] = []
                year = month_anchor.year
                month = month_anchor.month
                for _ in range(6):
                    offsets.append((year, month))
                    month -= 1
                    if month <= 0:
                        month = 12
                        year -= 1
                offsets.reverse()
                for year, month in offsets:
                    bucket_start = datetime(year, month, 1)
                    if month == 12:
                        next_month = datetime(year + 1, 1, 1)
                    else:
                        next_month = datetime(year, month + 1, 1)
                    bucket_end = next_month - timedelta(seconds=1)
                    trend_labels.append(bucket_start.strftime('%b'))
                    trend_ranges.append((bucket_start, bucket_end))

            trend_map: dict[int, dict[str, float]] = {fid: {label: 0.0 for label in trend_labels} for fid in fund_ids}
            earliest_dt = trend_ranges[0][0] if trend_ranges else now_dt

            donation_trend_rows = (
                db.session.query(Donation.fund_id, Donation.donation_date, Donation.amount)
                .filter(
                    Donation.organization_id == org.id,
                    Donation.fund_id.in_(fund_ids),
                    Donation.donation_date >= earliest_dt,
                )
                .all()
            )
            for d_fund_id, donation_date, amount in donation_trend_rows:
                if d_fund_id is None or donation_date is None:
                    continue
                fund_trend = trend_map.get(int(d_fund_id), {})
                for idx, (bucket_start, bucket_end) in enumerate(trend_ranges):
                    if bucket_start <= donation_date <= bucket_end:
                        label = trend_labels[idx]
                        if label in fund_trend:
                            fund_trend[label] += float(amount or 0.0)
                        break

            expense_trend_rows = (
                db.session.query(Expense.fund_id, Expense.paid_at, Expense.amount)
                .filter(
                    Expense.organization_id == org.id,
                    Expense.fund_id.in_(fund_ids),
                    Expense.paid_at >= earliest_dt,
                )
                .all()
            )
            for e_fund_id, paid_at, amount in expense_trend_rows:
                if e_fund_id is None or paid_at is None:
                    continue
                fund_trend = trend_map.get(int(e_fund_id), {})
                for idx, (bucket_start, bucket_end) in enumerate(trend_ranges):
                    if bucket_start <= paid_at <= bucket_end:
                        label = trend_labels[idx]
                        if label in fund_trend:
                            fund_trend[label] -= float(amount or 0.0)
                        break

            def _sparkline_points(values: list[float]) -> str:
                if not values:
                    return ''
                min_v = min(values)
                max_v = max(values)
                width = 72.0
                height = 22.0
                if max_v == min_v:
                    return ' '.join(f"{(idx * (width / max(1, len(values) - 1))):.2f},{height / 2:.2f}" for idx in range(len(values)))
                points: list[str] = []
                x_step = width / max(1, len(values) - 1)
                for idx, val in enumerate(values):
                    x = idx * x_step
                    y = height - ((val - min_v) / (max_v - min_v) * height)
                    points.append(f"{x:.2f},{y:.2f}")
                return ' '.join(points)

        total_in = 0.0
        total_out = 0.0
        for fund in funds:
            f_id = int(fund.id)
            donation_info = donations_by_fund.get(f_id, {'count': 0, 'amount': 0.0})
            expense_info = expenses_by_fund.get(f_id, {'count': 0, 'amount': 0.0})
            amount_in = float(donation_info['amount'])
            amount_out = float(expense_info['amount'])
            total_in += amount_in
            total_out += amount_out
            fund_metrics[f_id] = {
                'donation_count': int(donation_info['count']),
                'expense_count': int(expense_info['count']),
                'amount_in': round(amount_in, 2),
                'amount_out': round(amount_out, 2),
                'net': round(amount_in - amount_out, 2),
                'trend_labels': trend_labels,
                'trend_values': [round(trend_map.get(f_id, {}).get(label, 0.0), 2) for label in trend_labels],
                'sparkline_points': _sparkline_points([trend_map.get(f_id, {}).get(label, 0.0) for label in trend_labels]),
            }

        summary['total_in'] = round(total_in, 2)
        summary['total_out'] = round(total_out, 2)
        summary['net'] = round(total_in - total_out, 2)

        if has_activity in {'with', 'without'}:
            filtered_funds: list[Fund] = []
            for fund in funds:
                metrics = fund_metrics.get(int(fund.id), {})
                has_rows = int(metrics.get('donation_count', 0)) > 0 or int(metrics.get('expense_count', 0)) > 0
                if has_activity == 'with' and has_rows:
                    filtered_funds.append(fund)
                if has_activity == 'without' and not has_rows:
                    filtered_funds.append(fund)
            funds = filtered_funds

        def _fund_sort_key(item: Fund):
            metrics = fund_metrics.get(int(item.id), {})
            if sort_by == 'status':
                return 'active' if bool(getattr(item, 'is_active', False)) else 'inactive'
            if sort_by == 'updated':
                return getattr(item, 'updated_at', datetime.min)
            if sort_by == 'in':
                return float(metrics.get('amount_in', 0.0) or 0.0)
            if sort_by == 'out':
                return float(metrics.get('amount_out', 0.0) or 0.0)
            if sort_by == 'net':
                return float(metrics.get('net', 0.0) or 0.0)
            if sort_by == 'donations':
                return int(metrics.get('donation_count', 0) or 0)
            if sort_by == 'expenses':
                return int(metrics.get('expense_count', 0) or 0)
            return str(getattr(item, 'name', '') or '').strip().lower()

        funds.sort(key=_fund_sort_key, reverse=(sort_dir == 'desc'))

        filtered_total = len(funds)
        active_filtered_count = sum(1 for f in funds if bool(getattr(f, 'is_active', False)))
        total_pages = max(1, (filtered_total + per_page - 1) // per_page)
        if page > total_pages:
            page = total_pages
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        funds = funds[start_idx:end_idx]

        summary['total_funds'] = filtered_total
        summary['active_funds'] = active_filtered_count
        summary['inactive_funds'] = max(0, int(filtered_total) - int(active_filtered_count))
    else:
        total_pages = 1

    status_action_form = FundStatusActionForm()
    bulk_action_form = BulkFundActionForm()

    ai_context = {
        'active_page': 'funds',
        'organization': org.name if org else None,
        'fund_count': len(funds),
    }
    return render_template(
        'funds.html',
        funds=funds,
        fund_metrics=fund_metrics,
        summary=summary,
        status_action_form=status_action_form,
        bulk_action_form=bulk_action_form,
        active_page='funds',
        filter_q=query,
        filter_status=status,
        filter_has_activity=has_activity,
        filter_trend_window=trend_window,
        filter_sort_by=sort_by,
        filter_sort_dir=sort_dir,
        page=page,
        total_pages=total_pages,
        per_page=per_page,
        per_page_options=per_page_options,
        ai_context=ai_context,
    )


@main_bp.route('/funds/<int:fund_id>/status', methods=['POST'])
@login_required
@roles_required('admin', 'staff')
def fund_status_update(fund_id: int):
    org = _current_org()
    if not org:
        flash('No organization is available.', 'error')
        return redirect(url_for('main.funds_list'))

    form = FundStatusActionForm()
    if not form.validate_on_submit():
        flash('Invalid fund status request.', 'error')
        return redirect(url_for('main.funds_list'))

    if int(form.fund_id.data or 0) != int(fund_id):
        flash('Fund status request did not match record id.', 'error')
        return redirect(url_for('main.funds_list'))

    set_status = str(form.set_status.data or '').strip().lower()
    if set_status not in {'active', 'inactive'}:
        flash('Choose a valid fund status.', 'error')
        return redirect(url_for('main.funds_list'))

    try:
        FundService().update_fund(
            fund_id,
            org.id,
            actor_id=getattr(current_user, 'id', None),
            is_active=(set_status == 'active'),
        )
        flash(f'Fund status updated to {set_status}.', 'success')
    except FundNotFound:
        flash('Fund not found.', 'error')
    except ValueError as exc:
        flash(str(exc), 'error')
    except FundConcurrencyError:
        flash('This fund changed while updating. Please retry.', 'error')

    next_url = str(form.next_url.data or '').strip()
    if next_url.startswith('/'):
        return redirect(next_url)
    return redirect(url_for('main.funds_list'))


@main_bp.route('/funds/bulk/status', methods=['POST'])
@login_required
@roles_required('admin', 'staff')
def funds_bulk_status_update():
    org = _current_org()
    if not org:
        flash('No organization is available.', 'error')
        return redirect(url_for('main.funds_list'))

    form = BulkFundActionForm()
    if not form.validate_on_submit():
        flash('Invalid bulk fund status request.', 'error')
        return redirect(url_for('main.funds_list'))

    selected_ids = [
        int(raw_id)
        for raw_id in request.form.getlist('fund_ids')
        if str(raw_id or '').strip().isdigit()
    ]
    if not selected_ids:
        flash('Select at least one fund for bulk status update.', 'error')
        return redirect(url_for('main.funds_list'))

    set_status = str(form.set_status.data or '').strip().lower()
    if set_status not in {'active', 'inactive'}:
        flash('Choose a valid status to apply.', 'error')
        return redirect(url_for('main.funds_list'))

    updated = 0
    skipped = 0
    svc = FundService()
    for fund_id in selected_ids:
        try:
            svc.update_fund(
                fund_id,
                org.id,
                actor_id=getattr(current_user, 'id', None),
                is_active=(set_status == 'active'),
            )
            updated += 1
        except (FundNotFound, FundConcurrencyError, ValueError):
            skipped += 1

    flash(f'Bulk fund status update complete: {updated} updated, {skipped} skipped.', 'success' if updated else 'info')
    next_url = str(form.next_url.data or '').strip()
    if next_url.startswith('/'):
        return redirect(next_url)
    return redirect(url_for('main.funds_list'))


@main_bp.route('/funds/export/<string:file_type>')
@login_required
@require_step_up_auth
def funds_export(file_type: str):
    org = _current_org()
    if not org:
        flash('No organization available for export.', 'error')
        return redirect(url_for('main.funds_list'))

    query = request.args.get('q', '').strip()
    status = request.args.get('status', '').strip()

    funds = FundService().list_all_funds(org.id, search=query or None, status=status or None)
    headers = ['ID', 'Name', 'Description', 'Status', 'Updated At']
    rows = [
        [f.id, f.name, f.description or '', 'active' if f.is_active else 'inactive', f.updated_at.strftime('%Y-%m-%d %H:%M:%S') if f.updated_at else '']
        for f in funds
    ]

    if file_type == 'csv':
        data = _build_csv_bytes(headers, rows)
        return send_file(BytesIO(data), mimetype='text/csv', as_attachment=True, download_name='funds.csv')
    if file_type == 'xlsx':
        data = _build_xlsx_bytes('Funds', headers, rows)
        return send_file(
            BytesIO(data),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='funds.xlsx',
        )
    flash('Unsupported export type.', 'error')
    return redirect(url_for('main.funds_list'))


@main_bp.route('/funds/new', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'staff')
def fund_create():
    org = _current_org()
    if not org:
        flash('No organization is available. Please seed data first.', 'error')
        return redirect(url_for('main.dashboard'))

    form = FundForm()
    if form.validate_on_submit():
        try:
            FundService().create_fund(
                org.id,
                form.name.data,
                description=form.description.data,
                is_active=form.is_active.data == 'true',
                actor_id=getattr(current_user, 'id', None),
            )
        except ValueError as exc:
            flash(str(exc), 'error')
            return render_template('fund_form.html', form=form, is_edit=False, active_page='funds')
        except FundConcurrencyError:
            flash('This fund was changed by another user. Please reload and try again.', 'error')
            return render_template('fund_form.html', form=form, is_edit=False, active_page='funds')
        flash('Fund created successfully.', 'success')
        return redirect(url_for('main.funds_list'))

    return render_template('fund_form.html', form=form, is_edit=False, active_page='funds')


@main_bp.route('/funds/<int:fund_id>/edit', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'staff')
def fund_edit(fund_id: int):
    org = _current_org()
    try:
        fund = FundService().get_fund(fund_id, org.id)
    except FundNotFound:
        abort(404)
    form = FundForm(obj=fund)
    if request.method == 'GET':
        form.is_active.data = 'true' if fund.is_active else 'false'

    if form.validate_on_submit():
        try:
            FundService().update_fund(
                fund.id,
                org.id,
                actor_id=getattr(current_user, 'id', None),
                name=form.name.data,
                description=form.description.data,
                is_active=form.is_active.data == 'true',
            )
        except ValueError as exc:
            flash(str(exc), 'error')
            return render_template('fund_form.html', form=form, is_edit=True, fund=fund, active_page='funds')
        except FundConcurrencyError:
            flash('This fund was updated by another user. Please reload and submit again.', 'error')
            return render_template('fund_form.html', form=form, is_edit=True, fund=fund, active_page='funds')
        flash('Fund updated successfully.', 'success')
        return redirect(url_for('main.funds_list'))

    return render_template('fund_form.html', form=form, is_edit=True, fund=fund, active_page='funds')


@main_bp.route('/reports')
@login_required
def reports_page():
    org = _current_org()
    overview = {
        'total_donations': 0.0,
        'total_expenses': 0.0,
        'net_total': 0.0,
        'chart_data': {
            'labels': [],
            'donations': [],
            'expenses': [],
            'net': [],
            'totals': {'donations': 0.0, 'expenses': 0.0, 'net': 0.0},
        },
    }
    if org:
        overview = ReportingService().financial_overview(org.id)

    return render_template(
        'reports.html',
        total_donations=overview['total_donations'],
        total_expenses=overview['total_expenses'],
        net_total=overview['net_total'],
        chart_data=overview['chart_data'],
        active_page='reports',
        ai_context={
            'active_page': 'reports',
            'organization': org.name if org else None,
            'total_donations': overview['total_donations'],
            'total_expenses': overview['total_expenses'],
            'net_balance': overview['net_total'],
        },
    )


@main_bp.route('/campaigns/email-workbench')
@login_required
def campaign_email_workbench_page():
    org = _current_org()
    actor_role = str(getattr(current_user, 'role', '') or '').strip().lower()
    can_authorize_external_comms = bool(getattr(current_user, 'can_authorize_external_comms', False))
    can_send_campaign_email = actor_role == 'admin' or can_authorize_external_comms
    overview = {
        'total_donations': 0.0,
        'total_expenses': 0.0,
        'net_total': 0.0,
        'chart_data': {
            'labels': [],
            'donations': [],
            'expenses': [],
            'net': [],
            'totals': {'donations': 0.0, 'expenses': 0.0, 'net': 0.0},
        },
    }
    campaign_summary = {
        'total_campaigns': 0,
        'active_campaigns': 0,
    }
    if org:
        overview = ReportingService().financial_overview(org.id)
        campaigns = list(
            db.session.scalars(
                select(Campaign).where(Campaign.organization_id == org.id)
            )
        )
        campaign_summary['total_campaigns'] = len(campaigns)
        campaign_summary['active_campaigns'] = sum(1 for c in campaigns if str(getattr(c, 'status', '') or '') == 'active')

    return render_template(
        'campaigns/email_workbench.html',
        total_donations=overview['total_donations'],
        total_expenses=overview['total_expenses'],
        net_total=overview['net_total'],
        chart_data=overview['chart_data'],
        campaign_summary=campaign_summary,
        can_send_campaign_email=can_send_campaign_email,
        can_authorize_external_comms=can_authorize_external_comms,
        actor_role=actor_role,
        email_workbench_only=True,
        active_page='campaign_email_workbench',
        ai_context={
            'active_page': 'campaign_email_workbench',
            'organization': org.name if org else None,
            'total_donations': overview['total_donations'],
            'total_expenses': overview['total_expenses'],
            'net_balance': overview['net_total'],
        },
    )


@main_bp.route('/reports/compliance/evidence')
@login_required
@roles_required('admin', 'staff')
def reports_compliance_evidence():
    org = _current_org()
    scope = request.args.get('scope', 'org').strip().lower()
    if scope == 'global' and org is not None:
        return {'error': 'Global compliance scope is restricted to system administrators only.'}, 403
    organization_id = None if scope == 'global' else (org.id if org else None)

    payload = build_compliance_evidence(
        app=current_app,
        organization_id=organization_id,
    )

    as_download = request.args.get('download', '0').strip().lower() in {'1', 'true', 'yes', 'on'}
    if as_download:
        body = json.dumps(payload, indent=2, ensure_ascii=False).encode('utf-8')
        filename = f"compliance-evidence-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.json"
        return send_file(
            BytesIO(body),
            mimetype='application/json',
            as_attachment=True,
            download_name=filename,
        )
    return payload


@main_bp.route('/about')
def about():
    """About page."""
    return render_template('about.html', active_page='about')


@main_bp.route('/help')
def help():
    """Help/documentation page."""
    return render_template('help.html', active_page='help')


@main_bp.route('/calendar-deadlines')
@login_required
def calendar_deadlines_page():
    """Calendar and deadlines workspace."""
    return render_template('calendar_deadlines.html', active_page='calendar_deadlines')


@main_bp.route('/integrations-hub')
@login_required
@roles_required('admin', 'staff')
def integrations_hub_page():
    """Operational integrations workspace."""
    return render_template('integrations_hub.html', active_page='integrations_hub')


@main_bp.route('/ai-assistant-hub')
@login_required
def ai_assistant_hub_page():
    """AI assistant launcher and guardrails reference."""
    return render_template('ai_assistant_hub.html', active_page='ai_assistant_hub')


# ---------------------------------------------------------------------------
# Beneficiary Management UI
# ---------------------------------------------------------------------------

@main_bp.route('/beneficiaries')
@login_required
def beneficiaries_list():
    org = _current_org()
    query = request.args.get('q', '').strip()
    status_filter = request.args.get('status', '').strip()

    beneficiaries = []
    if org:
        all_beneficiaries = list_beneficiaries(org.id)
        if query:
            ql = query.lower()
            all_beneficiaries = [
                b for b in all_beneficiaries
                if ql in (b.first_name or '').lower()
                or ql in (b.last_name or '').lower()
                or ql in (b.email or '').lower()
                or ql in (b.program or '').lower()
            ]
        if status_filter:
            all_beneficiaries = [b for b in all_beneficiaries if b.status == status_filter]
        beneficiaries = all_beneficiaries

    ai_context = {
        'active_page': 'beneficiaries',
        'organization': org.name if org else None,
        'beneficiary_count': len(beneficiaries),
    }
    ctx = dict(
        beneficiaries=beneficiaries,
        active_page='beneficiaries',
        filter_q=query,
        filter_status=status_filter,
        ai_context=ai_context,
    )
    if request.headers.get('HX-Request'):
        return render_template('_beneficiaries_rows.html', **ctx)
    return render_template('beneficiaries.html', **ctx)


@main_bp.route('/beneficiaries/new', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'staff')
def beneficiary_new():
    org = _current_org()
    if not org:
        flash('No organization is available.', 'error')
        return redirect(url_for('main.beneficiaries_list'))

    form = BeneficiaryIntakeForm()
    if form.validate_on_submit():
        try:
            create_beneficiary(
                org.id,
                form.first_name.data,
                form.last_name.data or '',
                email=form.email.data or None,
                phone=form.phone.data or None,
                date_of_birth=form.date_of_birth.data,
                gender=form.gender.data or None,
                program=form.program.data or None,
                status=form.status.data,
                city=form.city.data or None,
                country=form.country.data or None,
                notes=form.notes.data or None,
            )
            flash('Beneficiary added successfully.', 'success')
            return redirect(url_for('main.beneficiaries_list'))
        except ValueError as exc:
            flash(str(exc), 'error')

    return render_template('beneficiary_form.html', form=form, is_edit=False, active_page='beneficiaries')


@main_bp.route('/beneficiaries/<int:beneficiary_id>')
@login_required
def beneficiary_detail(beneficiary_id: int):
    from ngo_homesuite.services.activity_timeline_service import ActivityTimelineService

    org = _current_org()
    if not org:
        abort(404)

    beneficiary = get_beneficiary(beneficiary_id, org.id)
    if not beneficiary:
        abort(404)

    cases = list_cases(org.id, beneficiary_id=beneficiary_id)

    activity_query = (request.args.get('activity_q') or '').strip()

    # Fetch unified activity timeline
    timeline_items = []
    try:
        timeline_items = ActivityTimelineService.get_beneficiary_timeline(
            org.id,
            beneficiary_id,
            limit=25,
            search_query=activity_query or None,
        )
    except Exception:
        # If timeline fetch fails, continue without it
        timeline_items = []

    beneficiary_ai_insights = None
    try:
        beneficiary_ai_insights = CopilotToolRegistry().execute(
            'summarize_activity_timeline',
            {
                'entity_type': 'beneficiary',
                'entity_id': beneficiary.id,
                'limit': 25,
                'query': activity_query or None,
            },
            {
                'organization_id': org.id,
                'actor': getattr(current_user, 'username', 'web'),
            },
        )
        if isinstance(beneficiary_ai_insights, dict) and beneficiary_ai_insights.get('error'):
            beneficiary_ai_insights = None
    except Exception:
        beneficiary_ai_insights = None

    ai_context = {
        'active_page': 'beneficiaries',
        'organization': org.name,
        'beneficiary_id': beneficiary.id,
        'beneficiary_name': f"{beneficiary.first_name} {beneficiary.last_name}".strip(),
        'case_count': len(cases),
    }
    return render_template(
        'beneficiary_detail.html',
        beneficiary=beneficiary,
        cases=cases,
        timeline_items=timeline_items,
        activity_query=activity_query,
        beneficiary_ai_insights=beneficiary_ai_insights,
        active_page='beneficiaries',
        ai_context=ai_context,
    )
